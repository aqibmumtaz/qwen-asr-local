import csv
import re
import subprocess
import wave
from collections import defaultdict
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
RESULTS = ROOT / "benchmark_results"
INPUT_CSV = RESULTS / "final_benchmark_sheet.csv"
OUT_XLSX = RESULTS / "final_benchmark_sheet.xlsx"
GPU_USAGE = "N/A (CPU-only, Metal OFF)"

MMPROJ = ROOT / "models/mmproj-Qwen3-ASR-1.7B-bf16-new.gguf"
LLAMA_MTMD = ROOT / "llama.cpp/build/bin/llama-mtmd-cli"
TRANSCRIBE_SCRIPT = ROOT / "transcribe_and_translate.sh"

PROMPT = "<|im_start|>user\\n<|audio_start|><|audio_pad|><|audio_end|><|im_end|>\\n<|im_start|>assistant\\n"
HINDI_HIGHLIGHT_MODELS = {
    "Qwen3-8B-Q5_K_M", "Qwen3-8B-Q8_0", "Qwen3-8B-Q4_K_M", "Qwen3-8B-BF16",
    "Qwen3-4B-Q8_0", "Qwen3-4B-Q4_K_M", "Qwen3-4B-BF16",
}


def audio_duration_seconds(path: Path) -> float:
    try:
        with wave.open(str(path), "rb") as wf:
            return wf.getnframes() / float(wf.getframerate())
    except Exception:
        return 0.0


def measure_peak_rss_bytes(shell_cmd: str) -> int:
    try:
        proc = subprocess.run(
            ["/usr/bin/time", "-l", "sh", "-c", shell_cmd],
            cwd=ROOT,
            capture_output=True,
            text=True,
            timeout=300,
        )
        text = (proc.stdout or "") + "\n" + (proc.stderr or "")
        match = re.search(
            r"^\s*(\d+)\s+maximum resident set size\s*$", text, flags=re.MULTILINE
        )
        return int(match.group(1)) if match else 0
    except Exception:
        return 0


def fmt_gb(b: int) -> str:
    return f"{(b / (1024 * 1024 * 1024)):.2f}"


def is_hindi_transcription(model: str, audio: str) -> bool:
    return model in HINDI_HIGHLIGHT_MODELS and audio in {
        "sample_ur1.wav",
        "sample_ur2.wav",
    }


try:
    with INPUT_CSV.open(encoding="utf-8-sig") as f:
        base_rows = list(csv.DictReader(f))
except Exception:
    print(f"Error reading {INPUT_CSV}")
    raise SystemExit(1)

if not base_rows:
    print("No data in base CSV")
    raise SystemExit(1)

sample_names = []
for row in base_rows:
    audio = row["audio"]
    if audio not in sample_names:
        sample_names.append(audio)

models = []
for row in base_rows:
    model = row["model"]
    if model not in models:
        models.append(model)

asr_models = []
for row in base_rows:
    asr = row["asr_model"]
    if asr not in asr_models:
        asr_models.append(asr)

audio_secs = {
    audio: audio_duration_seconds(ROOT / "samples" / audio) for audio in sample_names
}

print("Measuring peak RAM usage for translation models...")
rss_by_model = {}
for model in models:
    print(f"  Measuring {model}...")
    first_asr = asr_models[0]
    cmd = f"bash '{TRANSCRIBE_SCRIPT}' --asr-model '{first_asr}' --model '{model}' '{ROOT / 'samples' / sample_names[0]}' >/dev/null 2>&1"
    rss = measure_peak_rss_bytes(cmd)
    rss_by_model[model] = rss
    print(f"    {fmt_gb(rss)} GB")

print("Measuring ASR peak RAM for each decoder...")
asr_rss_by_model = {}
for asr_model in asr_models:
    print(f"  Measuring {asr_model}...")
    asr_model_path = ROOT / "models" / f"{asr_model}.gguf"
    asr_cmd = f"'{LLAMA_MTMD}' -m '{asr_model_path}' --mmproj '{MMPROJ}' --image '{ROOT / 'samples' / sample_names[0]}' -p '{PROMPT}' -n 256 --no-warmup >/dev/null 2>&1"
    asr_rss = measure_peak_rss_bytes(asr_cmd)
    asr_rss_by_model[asr_model] = asr_rss
    print(f"    {fmt_gb(asr_rss)} GB")

print("Building extended rows with all columns...")
extended_rows = []
for row in base_rows:
    audio = row["audio"]
    asr_model = row["asr_model"]
    asr_s = float(row.get("asr_seconds", 0))
    urdu_s = float(row.get("urdu_seconds", 0))
    total_s = asr_s + urdu_s
    audio_dur = audio_secs.get(audio, 1.0)
    sec_per_min = (total_s / audio_dur) * 60.0 if audio_dur > 0 else 0.0

    extended_rows.append(
        {
            "asr_model": asr_model,
            "model": row["model"],
            "audio": audio,
            "audio_seconds": f"{audio_dur:.2f}",
            "asr_seconds": f"{asr_s:.2f}",
            "translation_seconds_est": f"{urdu_s:.2f}",
            "total_inference_seconds": f"{total_s:.2f}",
            "total_seconds_per_min_audio": f"{sec_per_min:.2f}",
            "asr_peak_ram_gb": fmt_gb(asr_rss_by_model.get(asr_model, 0)),
            "pipeline_peak_ram_gb": fmt_gb(rss_by_model.get(row["model"], 0)),
            "gpu_usage": GPU_USAGE,
            "status": row["status"],
            "transcript": row.get("transcript", ""),
            "english_translation": row.get("english_translation", ""),
            "urdu_translation": row.get("urdu_translation", ""),
        }
    )


agg = defaultdict(
    lambda: {"total": 0.0, "asr": 0.0, "trans": 0.0, "spm": 0.0, "n": 0, "ok": 0}
)
for row in extended_rows:
    model = row["model"]
    agg[model]["total"] += float(row["total_inference_seconds"])
    agg[model]["asr"] += float(row["asr_seconds"])
    agg[model]["trans"] += float(row["translation_seconds_est"])
    agg[model]["spm"] += float(row["total_seconds_per_min_audio"])
    agg[model]["n"] += 1
    if row["status"] == "OK":
        agg[model]["ok"] += 1

model_summary = []
for model in models:
    n = agg[model]["n"] or 1
    model_summary.append(
        {
            "model": model,
            "avg_total": agg[model]["total"] / n,
            "avg_asr": agg[model]["asr"] / n,
            "avg_trans": agg[model]["trans"] / n,
            "avg_spm": agg[model]["spm"] / n,
            "ok": agg[model]["ok"],
            "fail": n - agg[model]["ok"],
            "ram_gb": fmt_gb(rss_by_model.get(model, 0)),
            "gpu_usage": GPU_USAGE,
        }
    )
model_summary.sort(key=lambda item: item["avg_total"])

print(f"Writing Excel report: {OUT_XLSX}")

HEADERS = [
    "ASR Model", "Translation Model", "Audio", "Audio Duration (s)", "ASR Time (s)",
    "Translation Time (s) est", "Total Inference (s)", "Processing Speed (s/min)",
    "ASR Peak RAM (GB)", "ASR + Trans Peak RAM (GB)",
    "GPU Usage", "Run Status", "ASR Transcript", "Translated English", "Translated Urdu",
]
FIELDS = [
    "asr_model", "model", "audio", "audio_seconds", "asr_seconds",
    "translation_seconds_est", "total_inference_seconds", "total_seconds_per_min_audio",
    "asr_peak_ram_gb", "pipeline_peak_ram_gb", "gpu_usage", "status", "transcript",
    "english_translation", "urdu_translation",
]

wb = Workbook()
ws = wb.active
ws.title = "Benchmark Results"

header_fill = PatternFill("solid", fgColor="2C3E50")
header_font = Font(bold=True, color="FFFFFF")
highlight_fill = PatternFill("solid", fgColor="FFF3CD")
alt_fill = PatternFill("solid", fgColor="F9F9F9")
info_font = Font(bold=True)
thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws.append(["ASR Models:", ", ".join(asr_models)])
ws["A1"].font = info_font
ws.append(["Translation Models:", ", ".join(models)])
ws["A2"].font = info_font
ws.append([])

ws.append(HEADERS)
for cell in ws[4]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
ws.row_dimensions[4].height = 24

for row in extended_rows:
    values = [row.get(f, "") for f in FIELDS]
    ws.append(values)
    excel_row = ws.max_row
    highlight = is_hindi_transcription(row["model"], row["audio"])
    fill = highlight_fill if highlight else (alt_fill if excel_row % 2 == 0 else None)
    for cell in ws[excel_row]:
        if fill:
            cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=False)

col_widths = [24, 22, 18, 14, 10, 18, 16, 20, 14, 22, 24, 8, 45, 45, 45]
for i, width in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = width

ws.freeze_panes = "A5"
wb.save(OUT_XLSX)

print("✓ DONE!")
print(f"  Excel    : {OUT_XLSX}")
for row in model_summary:
    print(
        f"{row['model']}: {row['avg_total']:.2f}s avg | {row['avg_trans']:.2f}s trans | {row['ram_gb']}GB RAM"
    )
