"""Quick preview: generates final_benchmark_sheet.xlsx from existing full CSV (no benchmarking runs)."""
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RESULTS = Path(__file__).resolve().parent
INPUT_CSV = RESULTS / "final_benchmark_full_sheet.csv"
OUT_XLSX = RESULTS / "final_benchmark_sheet.xlsx"

HINDI_HIGHLIGHT_MODELS = {
    "Qwen3-8B-Q5_K_M", "Qwen3-8B-Q8_0", "Qwen3-8B-Q4_K_M",
    "Qwen3-4B-Q8_0", "Qwen3-4B-Q4_K_M",
}
URDU_AUDIOS = {"sample_ur1.wav", "sample_ur2.wav"}

ASR_MODEL_NAME = "Qwen3-ASR-1.7B-Q8_0"

HEADERS = [
    "ASR Model", "Translation Model", "Audio", "Audio Duration (s)", "ASR Time (s)",
    "Translation Time (s) est", "Total Inference (s)", "Processing Speed (s/min)",
    "ASR Peak RAM (GB)", "ASR + Trans Peak RAM (GB)",
    "GPU Usage", "Run Status", "ASR Transcript", "Translated English", "Translated Urdu",
]
FIELDS = [
    "_asr_model", "model", "audio", "audio_seconds", "asr_seconds",
    "translation_seconds_est", "total_inference_seconds", "total_seconds_per_min_audio",
    "asr_peak_ram_gb", "pipeline_peak_ram_gb", "gpu_usage", "status", "transcript",
    "english_translation", "urdu_translation",
]

try:
    with INPUT_CSV.open(encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))
except Exception:
    print(f"Error reading {INPUT_CSV}")
    raise SystemExit(1)

trans_models = sorted({r["model"] for r in rows})

wb = Workbook()
ws = wb.active
ws.title = "Benchmark Results"

header_fill = PatternFill("solid", fgColor="2C3E50")
header_font = Font(bold=True, color="FFFFFF")
highlight_fill = PatternFill("solid", fgColor="FFF3CD")
alt_fill = PatternFill("solid", fgColor="F9F9F9")
info_font = Font(bold=True)
thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws.append(["ASR Model:", "Qwen3-ASR-1.7B-Q8_0"])
ws["A1"].font = info_font
ws.append(["Translation Model:", ", ".join(trans_models)])
ws["A2"].font = info_font
ws.append([])

ws.append(HEADERS)
for cell in ws[4]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
ws.row_dimensions[4].height = 24

for row in rows:
    row["_asr_model"] = ASR_MODEL_NAME
    values = [row.get(f, "") for f in FIELDS]
    ws.append(values)
    excel_row = ws.max_row
    highlight = row["model"] in HINDI_HIGHLIGHT_MODELS and row["audio"] in URDU_AUDIOS
    fill = highlight_fill if highlight else (alt_fill if excel_row % 2 == 0 else None)
    for cell in ws[excel_row]:
        if fill:
            cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=False)

col_widths = [24, 22, 18, 14, 10, 18, 16, 20, 14, 22, 24, 8, 45, 45, 45]
for i, width in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = width

ws.freeze_panes = "A5"

wb.save(OUT_XLSX)
print(f"✓ Written: {OUT_XLSX}")
