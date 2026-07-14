#!/usr/bin/env python3
"""
RIGOROUS validation of the resolver — old vs v2 vs v2+resolver.

WHY THIS EXISTS
    A previous test held out lexicon variants and reported "44% recovered". That
    number is real but it is a WEAK test, and the objection is fair: a held-out
    variant is by construction already close to its canonical, and it was collected
    from this very corpus. It shows the fuzzy matcher CAN work; it does not show
    that it helps on real ASR output.

    This script tests the thing that actually matters:
        take the ASR's REAL Roman output, find the words the exact lexicon CANNOT
        fix, and ask — does the resolver turn them into the word the human wrote?

THREE TESTS
    T1  LEAK CHECK          prove the held-out variants really were absent from the
                            exact map, and that the fuzzy index contains ONLY
                            canonicals (so holding out variants cannot leak).
    T2  REAL ASR WORDS      the honest end-to-end test. For every ASR word that the
                            exact lexicon misses, compare against the aligned gold
                            word. Count: FIXED / BROKEN / no-change.
    T3  THREE-WAY WER       old vs v2 vs v2+resolver, full pipeline, on all turns.

Usage:
    python3 validate_resolver.py
"""

import json
import random
import sys
from collections import Counter
from difflib import SequenceMatcher
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR))

import openpyxl
from resolver import Resolver, load_gold_vocab, normalise

V2 = SCRIPT_DIR / "data" / "lexicons_v2.json"
XLSX = SCRIPT_DIR / "data" / "CLL analysis" / "turnwise_results_eval_full.xlsx"


def turns():
    wb = openpyxl.load_workbook(str(XLSX), data_only=True)
    rs = list(wb["asr_results"].iter_rows(values_only=True))
    idx = {h: i for i, h in enumerate(rs[0])}
    out = []
    for r in rs[1:]:
        m = r[idx["roman_urdu_model"]]
        g = r[idx["roman_urdu_reference"]]
        if isinstance(m, str) and isinstance(g, str) and m.strip() and g.strip():
            out.append((m, g))
    return out


# ── T1: prove there is no leak ───────────────────────────────────────────────
def t1_leak_check():
    print("=" * 78)
    print("  T1 — LEAK CHECK: was the held-out test honest?")
    print("=" * 78)

    raw = json.loads(V2.read_text(encoding="utf-8"))["lexicons"]
    rng = random.Random(42)
    train = {"lexicon": {}, "phrases": raw["phrases"]}
    heldout = []
    for canon, variants in raw["lexicon"].items():
        kept = []
        for v in variants:
            if rng.random() < 0.5 and len(v) >= 6:
                heldout.append((v, canon))
            else:
                kept.append(v)
        train["lexicon"][canon] = kept

    tmp = Path("/tmp/_v2_leakcheck.json")
    tmp.write_text(json.dumps({"lexicons": train}), encoding="utf-8")
    r = Resolver(v2_path=tmp, gold_vocab=None, min_len=8)

    still_exact = [v for v, _ in heldout if v.lower() in r.exact]
    print(f"\n  held-out variants                 : {len(heldout)}")
    print(f"  still reachable by EXACT lookup   : {len(still_exact)}   "
          f"{'<-- LEAK!' if still_exact else '(none — clean)'}")

    # the fuzzy index must contain canonicals ONLY
    all_variants = {v.lower() for vs in raw["lexicon"].values() for v in vs}
    indexed = {c for cs in r.norm_exact.values() for c in cs}
    leaked = indexed & all_variants - {c.lower() for c in raw["lexicon"]}
    print(f"  fuzzy index size                  : {len(indexed)} canonicals")
    print(f"  variants leaked into fuzzy index  : {len(leaked)}   "
          f"{'<-- LEAK!' if leaked else '(none — index is canonicals only)'}")

    ok = not still_exact and not leaked
    print(f"\n  => the held-out test was {'HONEST ✓' if ok else 'CONTAMINATED ✗'}")
    print("     (the resolver had to compute each match from the canonical alone)")
    print()
    return ok


# ── T2: the honest test — real ASR words the lexicon cannot fix ──────────────
def align(model: str, gold: str):
    """Word-align ASR output to gold so we can judge each substitution."""
    m, g = model.split(), gold.split()
    sm = SequenceMatcher(a=[w.lower() for w in m], b=[w.lower() for w in g])
    pairs = []
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag == "replace" and (i2 - i1) == (j2 - j1):
            for k in range(i2 - i1):
                pairs.append((m[i1 + k], g[j1 + k]))
    return pairs


def t2_real_asr(min_len: int = 8):
    print("=" * 78)
    print("  T2 — THE HONEST TEST: real ASR words the EXACT lexicon cannot fix")
    print("=" * 78)
    print("      For each such word we know what the human actually wrote.")
    print("      FIXED  = resolver produced the gold word")
    print("      BROKEN = resolver changed it to something that is NOT the gold word")
    print()

    gold_vocab = load_gold_vocab()
    r = Resolver(gold_vocab=gold_vocab, min_len=min_len)

    fixed, broken, unchanged = Counter(), Counter(), 0
    considered = 0
    for model, gold in turns():
        for mw, gw in align(model, gold):
            lw = mw.lower()
            if lw in r.exact:
                continue                      # the exact lexicon handles it
            if lw == gw.lower():
                continue                      # already correct
            considered += 1
            out = r.resolve_word(mw)
            if out.lower() == gw.lower():
                fixed[f"{mw} -> {out}"] += 1
            elif out.lower() != lw:
                broken[f"{mw} -> {out}  (gold: {gw})"] += 1
            else:
                unchanged += 1

    nf, nb = sum(fixed.values()), sum(broken.values())
    print(f"  ASR words the exact lexicon MISSED and that are WRONG : {considered}")
    print()
    print(f"    FIXED by resolver     : {nf:>4}")
    print(f"    BROKEN by resolver    : {nb:>4}   (changed, but not to the gold word)")
    print(f"    left unchanged        : {unchanged:>4}")
    print(f"    NET                   : {nf-nb:>+4}")
    print()
    if fixed:
        print("    FIXED:")
        for k, n in fixed.most_common(12):
            print(f"      {n:>2}x  {k}")
    if broken:
        print("\n    BROKEN:")
        for k, n in broken.most_common(12):
            print(f"      {n:>2}x  {k}")
    print()
    return nf, nb


# ── T3: three-way WER, full pipeline ─────────────────────────────────────────
def wer_acc(hyp, ref):
    h, r = hyp.lower().split(), ref.lower().split()
    if not r:
        return 1.0
    dp = list(range(len(h) + 1))
    for rw in r:
        nd = [dp[0] + 1] + [0] * len(h)
        for j, hw in enumerate(h):
            nd[j + 1] = min(dp[j] + (0 if hw == rw else 1), dp[j + 1] + 1, nd[j] + 1)
        dp = nd
    return max(0.0, 1.0 - dp[len(h)] / len(r))


def t3_three_way():
    import importlib
    import os

    print("=" * 78)
    print("  T3 — THREE-WAY WER on the full pipeline")
    print("=" * 78)

    wb = openpyxl.load_workbook(str(XLSX), data_only=True)
    rs = list(wb["asr_results"].iter_rows(values_only=True))
    idx = {h: i for i, h in enumerate(rs[0])}
    data = [(r[idx["model_output_hindi"]], r[idx["roman_urdu_reference"]])
            for r in rs[1:]
            if isinstance(r[idx["model_output_hindi"]], str)
            and isinstance(r[idx["roman_urdu_reference"]], str)]

    res = {}
    for cfg in ("old", "v2", "v2+resolver"):
        os.environ["LEXICON"] = "old" if cfg == "old" else "v2"
        os.environ["RESOLVER"] = "1" if "resolver" in cfg else "0"
        import hindi_to_roman_urdu as H
        importlib.reload(H)
        accs = [wer_acc(H.transliterate(h), g) for h, g in data]
        res[cfg] = sum(accs) / len(accs)

    print()
    print(f"  {'config':<16} {'WER accuracy':>14}   {'vs old':>9}")
    print(f"  {'-'*16} {'-'*14}   {'-'*9}")
    for cfg in ("old", "v2", "v2+resolver"):
        d = (res[cfg] - res["old"]) * 100
        print(f"  {cfg:<16} {res[cfg]*100:>13.2f}%   {d:>+8.2f}%")
    print()
    print(f"  resolver adds over v2: {(res['v2+resolver']-res['v2'])*100:+.2f} pts")
    print()
    return res


if __name__ == "__main__":
    honest = t1_leak_check()
    nf, nb = t2_real_asr()
    res = t3_three_way()

    print("=" * 78)
    print("  VERDICT")
    print("=" * 78)
    print(f"  held-out test was honest      : {'YES' if honest else 'NO'}")
    print(f"  on REAL ASR words the lexicon misses: {nf} fixed, {nb} broken "
          f"(net {nf-nb:+d})")
    print(f"  end-to-end WER: old {res['old']*100:.2f}%  ->  v2 {res['v2']*100:.2f}%"
          f"  ->  +resolver {res['v2+resolver']*100:.2f}%")
    print()
    print("  The lexicon (v2) is where the win is. The resolver is insurance for")
    print("  spellings not yet enumerated — it barely moves THIS eval set because")
    print("  v2's variants were harvested from THIS corpus.")
    print()
