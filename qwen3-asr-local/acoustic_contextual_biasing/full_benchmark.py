"""
FULL, RESUMABLE benchmark over all 80 calls — remote GPU Qwen3-ASR.

Compares, per call (concatenate chunks -> transliterate() lexicon layer -> vs benchmark):
  prev      previous model (vendor's model_output_hindi)
  en        remote re-ASR, /en variant        (no domain bias)
  chughtai  remote re-ASR, /chughtai variant  (DOMAIN-BIASED for Chughtai Lab)

en vs chughtai = the contextual-biasing effect. prev vs chughtai = new stack vs old.

Transcription is CACHED per (chunk, variant) to transcription_cache.jsonl and resumable
(re-running skips done work; kill/restart any time). Reporting is fast and works on the
partial cache.

  python -m acoustic_contextual_biasing.full_benchmark --transcribe --variants chughtai,en
  python -m acoustic_contextual_biasing.full_benchmark --report
"""
from __future__ import annotations

import argparse
import importlib
import json
import os
import re
import sys
import time
import warnings
from collections import defaultdict
from difflib import SequenceMatcher
from pathlib import Path

warnings.filterwarnings("ignore")
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT)); sys.path.insert(0, str(ROOT / "benchmark"))

import openpyxl
from test_accuracy import diff_words

from .asr import RemoteASR

XLSX = ROOT / "benchmark" / "lab_test_80_calls_urdu_roman_urdu.xlsx"
AUDIO = ROOT / "benchmark" / "lab_test_80_audios_chunks_25s"
CACHE = Path(__file__).resolve().parent / "transcription_cache.jsonl"
CAP = re.compile(r"\b[A-Z][a-z]{2,}\b")


def load_calls():
    wb = openpyxl.load_workbook(str(XLSX)); ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True)); idx = {h: i for i, h in enumerate(rows[0])}
    calls = defaultdict(list)
    for r in rows[1:]:
        calls[r[idx["call_id"]]].append(r)
    for cid in calls:
        calls[cid].sort(key=lambda x: x[idx["chunk_index"]] or 0)
    return calls, idx


def read_cache() -> dict:
    out = {}
    if CACHE.exists():
        for line in CACHE.read_text(encoding="utf-8").splitlines():
            if line.strip():
                r = json.loads(line)
                out.setdefault(r["key"], {}).update(r.get("variants", {}))
    return out


def append_cache(key: str, variants: dict):
    with CACHE.open("a", encoding="utf-8") as f:
        f.write(json.dumps({"key": key, "variants": variants}, ensure_ascii=False) + "\n")


def transcribe(args):
    variants = args.variants.split(",")
    asr = {v: RemoteASR(variant=v) for v in variants}
    calls, _ = load_calls()
    cache = read_cache()

    todo = []
    for cid in calls:
        if not (AUDIO / str(cid)).exists():
            continue
        for ch in sorted((AUDIO / str(cid)).glob("chunk_*.wav")):
            key = f"{cid}/{ch.name}"
            missing = [v for v in variants if v not in cache.get(key, {})]
            if missing:
                todo.append((key, ch, missing))
    print(f"chunk-variants to transcribe: {sum(len(m) for _,_,m in todo)}  "
          f"({len(todo)} chunks; cache has {len(cache)})", flush=True)

    for n, (key, ch, missing) in enumerate(todo, 1):
        got = {}
        t0 = time.time()
        for v in missing:
            try:
                got[v] = asr[v].transcribe(ch)
            except Exception as e:
                print(f"  ! {key} [{v}] error: {type(e).__name__} {str(e)[:80]}", flush=True)
        if got:
            append_cache(key, got)
        print(f"  [{n}/{len(todo)}] {key} {list(got)} {time.time()-t0:.0f}s", flush=True)
    print("transcription pass complete.", flush=True)


def fuzzy_in(name, words, thr=0.75):
    n = name.lower()
    return any(SequenceMatcher(a=n, b=w, autojunk=False).ratio() >= thr for w in words)


def wer_acc(benchmark: str, hypothesis: str) -> tuple:
    """Classic word edit-distance. Returns (edits, ref_len)."""
    from test_accuracy import normalize_tokens
    ref = normalize_tokens(benchmark)
    hyp = normalize_tokens(hypothesis)
    if not ref:
        return (0, 0)
    dp = list(range(len(hyp) + 1))
    for rw in ref:
        nd = [dp[0] + 1] + [0] * len(hyp)
        for j, hw in enumerate(hyp):
            nd[j + 1] = min(dp[j] + (0 if hw == rw else 1), dp[j + 1] + 1, nd[j] + 1)
        dp = nd
    return (dp[len(hyp)], len(ref))


def report(args):
    os.environ["LEXICON"] = "v2"; os.environ["RESOLVER"] = "0"; os.environ["PHONETIC"] = ""
    import hindi_to_roman_urdu as H
    importlib.reload(H)
    calls, idx = load_calls()
    cache = read_cache()
    variants = args.variants.split(",")
    cfgs = ["prev"] + variants
    agg = {c: [0, 0] for c in cfgs}
    nm = {c: [0, 0] for c in cfgs}
    n_calls = n_full = 0

    # per-call results for Excel
    per_call_results = {}
    # per-chunk results for the detail sheet
    all_chunk_rows = []

    for cid, cr in calls.items():
        bench = next((x[idx["benchmark_roman_urdu"]] for x in cr
                      if isinstance(x[idx["benchmark_roman_urdu"]], str)
                      and x[idx["benchmark_roman_urdu"]].strip()), None)
        actual_urdu = next((x[idx["actual_urdu_transcript"]] for x in cr
                            if isinstance(x[idx.get("actual_urdu_transcript", -1)], str)
                            and x[idx["actual_urdu_transcript"]].strip()), "") if "actual_urdu_transcript" in idx else ""
        if not bench or not (AUDIO / str(cid)).exists():
            continue
        n_calls += 1
        chunks = sorted((AUDIO / str(cid)).glob("chunk_*.wav"))
        keys = [f"{cid}/{ch.name}" for ch in chunks]
        if not all(all(v in cache.get(k, {}) for v in variants) for k in keys):
            continue
        n_full += 1
        gnames = sorted(set(CAP.findall(bench)))
        texts = {"prev": " ".join(str(x[idx["model_output_hindi"]]).strip() for x in cr
                                  if isinstance(x[idx["model_output_hindi"]], str)
                                  and x[idx["model_output_hindi"]].strip())}
        for v in variants:
            texts[v] = " ".join(cache[k][v] for k in keys)

        # per-chunk detail rows
        for i, (ch, key) in enumerate(zip(chunks, keys)):
            row = {
                "audio_name": ch.name,
                "call_id": cid,
                "chunk_index": i,
                "actual_urdu_transcript": actual_urdu if i == 0 else "",
                "benchmark_roman_urdu": bench if i == 0 else "",
            }
            for v in variants:
                row[f"model_output_{v}"] = cache.get(key, {}).get(v, "")
            all_chunk_rows.append(row)

        d_call = {"bench": bench}
        for c in cfgs:
            roman = H.transliterate(texts[c])
            d = diff_words(bench, roman)
            agg[c][0] += d.matched; agg[c][1] += d.total
            words = roman.lower().split()
            nm[c][0] += sum(1 for n in gnames if fuzzy_in(n, words)); nm[c][1] += len(gnames)
            d_call[c] = {"roman": roman, "acc": d, "wer": wer_acc(bench, roman)}
        per_call_results[cid] = d_call

    label = {"prev": "previous model (vendor)", "en": "remote /en (no bias)",
             "chughtai": "remote /chughtai (biased)"}
    print("=" * 72)
    print(f"  FULL BENCHMARK — remote GPU ASR — {n_full}/{n_calls} calls fully transcribed")
    print("=" * 72)
    print(f"  {'config':<32}{'diff_words':>11}{'name recovery':>16}")
    print("  " + "-" * 58)
    for c in cfgs:
        if agg[c][1] == 0:
            continue
        dw = 100 * agg[c][0] / agg[c][1]
        nr = 100 * nm[c][0] / max(nm[c][1], 1)
        print(f"  {label.get(c, c):<32}{dw:>10.2f}%{nr:>13.1f}%  ({nm[c][0]}/{nm[c][1]})")
    print()

    # ---- Write to Excel ----
    if not per_call_results:
        print("  (no complete calls to write to Excel)")
        return

    OUT = ROOT / "benchmark" / "lab_test_80_calls_urdu_roman_urdu_benchmarked.xlsx"
    wb_out = openpyxl.load_workbook(str(OUT))

    # SHEET: model_biased_chughtai — per-chunk output
    sheet_name = "model_biased_chughtai"
    if sheet_name in wb_out.sheetnames:
        del wb_out[sheet_name]
    s2 = wb_out.create_sheet(sheet_name)
    headers = ["audio_name", "call_id", "chunk_index",
               "actual_urdu_transcript", "benchmark_roman_urdu"]
    for v in variants:
        headers.append(f"model_output_{v}")
    s2.append(headers)
    for row in all_chunk_rows:
        s2.append([row.get(h, "") for h in headers])

    # SHEET: benchmark_biasing_summary — per-call scores
    sum_name = "benchmark_biasing_summary"
    if sum_name in wb_out.sheetnames:
        del wb_out[sum_name]
    s3 = wb_out.create_sheet(sum_name)
    sum_headers = ["call_id", "benchmark_roman_urdu"]
    for c in cfgs:
        sum_headers += [f"out_{c}", f"acc_{c}", f"wer_{c}"]
    s3.append(sum_headers)

    def werp(e, n):
        return round(100 * max(0.0, 1 - e / n), 2) if n else None

    for cid, d in per_call_results.items():
        row = [cid, d["bench"]]
        for c in cfgs:
            row.append(d[c]["roman"])
            row.append(round(d[c]["acc"].accuracy, 4))
            row.append(werp(*d[c]["wer"]))
        s3.append(row)

    # corpus totals
    s3.append([])
    corpus_row = ["CORPUS (diff_words %)"]
    corpus_row.append("")
    for c in cfgs:
        dw = 100 * agg[c][0] / agg[c][1] if agg[c][1] else 0
        corpus_row += [round(dw, 2), "", ""]
    s3.append(corpus_row)

    wb_out.save(str(OUT))
    print(f"  written to: {OUT.name}")
    print(f"    - {sheet_name}          (per-chunk biased output)")
    print(f"    - {sum_name}   (per-call accuracy scores)")
    print()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--transcribe", action="store_true")
    ap.add_argument("--report", action="store_true")
    ap.add_argument("--variants", default="chughtai,en")
    args = ap.parse_args()
    if args.transcribe:
        transcribe(args)
    if args.report or not args.transcribe:
        report(args)


if __name__ == "__main__":
    main()
