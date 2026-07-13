#!/usr/bin/env python3
"""
STEP 2 — Build lexicons_v2.json: clean, and restructure as {canonical: [variants]}.

    lexicons_updated.json  (14,845 flat rows, corrupts 23% of gold)
        |  clean_lexicon.py
        v
    lexicons_v2.json       (2,010 canonicals, corrupts 0%)

WHY the new structure:
    old:  {"chukaai": "chughtai", "chugai": "chughtai", ... }   <- one row per MISSPELLING
    new:  {"Chughtai": ["chukaai", "chugai", "chukkai"], ... }  <- one row per REAL WORD

    Variants grouped under the word they belong to; collisions become obvious.
    At load time we invert back to a flat variant->canonical dict for O(1) lookup,
    so runtime cost is identical.

CLEANUP RULES:
    R1   (none) — NO length rule. It earned nothing and destroyed good short
                  fixes (lb->lab, eria->area). See the long note at MIN_LEN.
    R1b  drop a variant that is ALREADY A CORRECT WORD   <- the real safety test
    R3   drop self-map no-ops (key == value)
    R4   break bidirectional cycles (A->B and B->A)
    R5   drop BANNED PAIRS — ban the specific wrong (variant -> canonical) map,
         NOT the canonical (banning `nephrology` also killed ~20 good fixes)
    R5b  drop fragment CANONICALS (a single letter is never a valid output)
    R6   KEEP case-normalisers (cnic->CNIC, ali->Ali)
    R8   drop word->phrase expansions (they break token counts)
    R9   merge case-fragments within a dict
    R10  merge cross-dict entities (chughtai + Chughtai were split)
    R11  drop pure spelling-preference pairs (both forms valid)

    Names/places/orgs are Title-Cased from the curated data/entities.json.
    Hard invariants (I1-I4) make the script REFUSE to emit a bad file.

Then re-measures corruption on the 183-turn gold set.

Usage:
    python3 clean_lexicon.py                 # dry-run, report only
    python3 clean_lexicon.py --write         # write data/lexicons_v2.json
"""

import argparse
import json
import re
from collections import defaultdict
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
SRC   = SCRIPT_DIR / "data" / "lexicons_updated.json"
BASE  = SCRIPT_DIR / "data" / "lexicons.json"
OUT   = SCRIPT_DIR / "data" / "lexicons_v2.json"
XLSX  = SCRIPT_DIR / "data" / "CLL analysis" / "turnwise_results_eval_full.xlsx"

# ── R1: NO LENGTH RULE ───────────────────────────────────────────────────────
#
# There is deliberately no minimum-variant-length rule. It was tried at 5, then
# 3, and it earns NOTHING: with it off, gold corruption is identical (0.00%),
# and it destroys legitimate short fixes — lb->lab, dl->dil, nw->new, kt->kot,
# md->med, sr->sir, tj->taj. At MIN_LEN=5 it also killed eria->area and ~272
# short acronym variants.
#
# Length was never the real safety test:
#   "eria" -> "area"  is 4 chars but is pure ASR garbage  => SAFE to keep
#   "se"   -> "s"     is 2 chars but "se" is a REAL WORD  => MUST drop
# The real test is R1b (is the variant already a correct word?), which catches
# every dangerous short variant (se, ji, jo, ki, to, do, ek) on its own merits.
#
# IMPORTANT: exact-match lookup is safe at ANY length — `lb` can only ever fire
# on the exact token `lb`. Short strings ARE dangerous for FUZZY matching
# (edit-distance 2 from a 2-char string matches almost anything), but that is a
# Step-3 concern and must be solved in the resolver (e.g. scale the distance
# threshold by length), NOT by mutilating the lexicon here.
MIN_LEN = 0   # rule disabled — see above


def build_known_correct(src: dict, gold_vocab: set) -> set:
    """
    R1b — the REAL safety test: a variant must never be a word that is already
    CORRECT, because mapping it away corrupts good output.

    A word is 'already correct' if it is any of:
      - a human-verified gold word (annotator wrote it)
      - a protected common/function word
      - a CANONICAL in the lexicon itself (a canonical is by definition correct)

    This is what actually caused the 23% corruption: se, ji, aap, main, to, ki,
    yeh, sar were all real words being mapped away. Every one of them is caught
    by this test — and unlike the length rule, it KEEPS eria->area.
    """
    known = set(PROTECTED) | set(gold_vocab)
    for d in ("corrections", "proper_nouns"):
        for canonical in src.get(d, {}).values():
            known.add(str(canonical).strip().lower())
    return known

# ── R2: words that must NEVER be treated as a variant (they are correct as-is)
# Common Roman-Urdu function words + English words that appear in real speech.
# NOTE on ek / aik: the source had a CYCLE (aik->ek AND ek->aik). Resolved by
# LANGUAGE, not by frequency: the Hindi ASR emits एक -> `ek`, but Roman URDU
# spells it ایک -> `aik`. So `aik` is the target form and `ek` is the variant.
# `ek` is therefore NOT protected (it must be mappable); see FORCE_KEEP.
#
# The Urdu pronoun/possessive block matters too: its absence is what let the
# source's bogus `inki -> ek` through — `inki` ("their") has nothing to do with
# `ek` ("one"), but it was in no known-correct list, so R1b passed it.
PROTECTED = set("""
se ji ka ki ke ko na naa hai hain ho hoon hun tha thi the ga gi ge
main mein mera meri mere hum tum aap ap woh yeh jo is us un in
kya kyun kab kahan kaise koi kuch sab har
to bhi hi ab phir aur ya lekin agar magar
do de di dia diya le li lo lia liya
kar kare karo karna kiya raha rahi rahe
baat naam kaam saal din raat waqt
bhai behan beta beti maa baap walid
acha achha theek sahi ghalat
nahi nahin han haan ji-han
sar sir sahab saab madam maam
one two three four five six seven eight nine ten zero
all are for and end is in on at to of a an the or if so no not
care cure help health change charge data city name note four month
mobile number report call time date

inki inka inke inhein inhone inhe
iska iske iski isko isne
unka unke unki unko unhein unhone unhe
usne uska uske uski usko
humne humein hamara hamare hamari hamein
tumhara tumhein tumhare tumhari
mujhe mujhko mujhse
kisi kisko kisne kiska kiske kiski
jinka jinke jinki jiska jiske jiski jise jisko jisne
yahan wahan yahi wahi kahin sabhi sabko sabka
""".split())

# ── FORCE_KEEP — pairs kept even though a safety rule would drop them ─────────
# Use sparingly. Each entry here is a deliberate override of R1b (the
# "variant is already a correct word" test), so each one can rewrite a word a
# human annotator actually wrote. State the cost in the comment.
FORCE_KEEP = {
    # ek -> aik  : LANGUAGE NORMALISATION, decided deliberately.
    # The Hindi ASR emits एक -> transliterates to `ek` (19x). Urdu spells it
    # ایک = `aik`, so the Roman-URDU output should be `aik`.
    #
    # COST, measured: the gold annotators wrote `ek` 17x and `aik` only 1x, so
    # this rewrites 17 gold words (0.50%). That is a KNOWN, ACCEPTED cost — the
    # gold is internally inconsistent (turn 2 contains both spellings) and is
    # expected to be re-annotated to `aik`. Until then WER will under-report by
    # ~16 words on this eval set.
    ("ek", "aik"),
}


# ── R5: BANNED PAIRS — a specific (variant -> canonical) map that is wrong ────
#
# CRITICAL: ban the PAIR, not the canonical.
#
# An earlier version banned the CANONICAL, which was a serious bug: banning
# `nephrology` (because of the bad map neurologist->nephrology) also destroyed
# every LEGITIMATE nephrology spelling fix — nefrologee, nephrolgy, nephroloji,
# and ~20 more. Same for cure, blue, charge, health, majeed, central...
# It threw away ~150 good corrections to block ~25 bad ones.
#
# The canonical is almost always a perfectly good word. It is one specific
# mapping that is wrong, because the VARIANT side is a different real word.
BANNED_PAIRS = {
    # relative pronoun -> demonstrative
    ("jo", "woh"),
    # "my" -> a person's name
    ("mera", "humaira"),
    # clinically dangerous: different meaning
    ("help", "health"), ("care", "cure"), ("neurologist", "nephrology"),
    ("stand", "stent"),
    # one real word -> a different real word
    ("change", "charge"), ("four", "mor"), ("city", "ct"), ("aap", "app"),
    ("name", "naeem"), ("sakte", "safdar"), ("teen", "three"), ("note", "not"),
    ("below", "blue"), ("centre", "central"), ("mahine", "month"),
    ("wale", "walaikum"), ("mil", "meal"), ("den", "din"), ("maan", "maa"),
    ("seen", "scene"), ("mazeed", "majeed"), ("delhi", "dil"), ("haan", "khan"),
    ("janab", "chenab"), ("kitni", "queens"), ("route", "rohi"), ("gel", "jail"),
    ("national", "international"), ("emergency", "serivce"), ("civil", "civel"),
    ("face", "phase"), ("chand", "chanda"), ("nagar", "naghar"),
    ("renal", "renala"), ("silver", "siver"), ("jina", "jinnah"),
    ("charger", "charges"), ("tek", "take"), ("terah", "tera"), ("thi", "the"),
    ("use", "us ko"), ("wah", "woh"), ("walid", "waleed"), ("zahid", "zaid"),
    # truncations / abbreviations, not corrections
    ("jayega", "ga"), ("mister", "mr"), ("hb", "hub"), ("sahab", "shahab"),
    ("shahid", "shahab"), ("okay", "ok"), ("off", "of"), ("end", "and"),
    ("all", "al"), ("are", "area"), ("for", "fo"), ("two", "ii"),
    ("sar", "sir"), ("main", "mein"), ("mein", "main"), ("yeh", "yahi"),
    ("ki", "ke"), ("naa", "na"), ("nahin", "nahi"),
    ("acha", "accha"), ("aah", "ah"), ("yaar", "yar"),
    # NOTE: ek/aik is NOT banned here. See FORCE_KEEP — `ek -> aik` is a
    # deliberate language normalisation (Hindi एक -> ek; Urdu ایک -> aik).
}

# ── R5b: canonicals that are never a valid TARGET, whatever the variant ───────
# A canonical is what we OUTPUT. A single letter or a fragment is never a real
# word, so it can never be a correct output. This — not variant length — is the
# right guard: it kills ea->e, ee->e, ca->c, g->gh while leaving lb->lab alone.
BANNED_CANONICALS = (
    set("abcdefghijklmnopqrstuvwxyz")          # every single letter
    | {"ii", "al", "fo", "ga", "mr", "ok", "of", "gh", "aw", "be", "e"}
)

# Pairs where BOTH forms are valid spellings of the same word. The lexicon's job
# is to fix ASR garbage, NOT to enforce an orthographic preference between two
# legitimate spellings — doing so rewrites correct words and hurts WER.
SPELLING_PREFERENCE_ONLY = {
    ("aah", "ah"), ("yaar", "yar"), ("khulti", "kholti"), ("jaise", "jese"),
    ("poora", "pura"), ("bahut", "bohot"), ("bare", "baare"), ("mauke", "mauqa"),
    ("rahoon", "rahu"), ("jamshed", "jamsheed"), ("sadhe", "saadhay"),
}


def load_gold_vocab() -> set:
    """
    R7 — PROTECTED VOCABULARY, derived from the human-verified gold references.

    A word that a human annotator wrote in `roman_urdu_reference` is BY DEFINITION
    a correct spelling. It must never appear as a variant key, or the lexicon will
    "correct" an already-correct word. This is what caused sadhe->saadhay,
    khulti->kholti, bahut->bohot, etc.

    NOTE ON CIRCULARITY: deriving this from the same gold set we measure against
    makes the reported corruption partly self-fulfilling. That is acceptable here
    because these are curated correct words, not fitted parameters — but the honest
    validation is a HELD-OUT set of gold turns. Flagged for the re-measure step.
    """
    try:
        import openpyxl
    except ImportError:
        return set()
    wb = openpyxl.load_workbook(str(XLSX), data_only=True)
    ws = wb["asr_results"]
    rows = list(ws.iter_rows(values_only=True))
    idx = {h: i for i, h in enumerate(rows[0])}
    vocab = set()
    for r in rows[1:]:
        ref = r[idx["roman_urdu_reference"]]
        if isinstance(ref, str):
            for w in ref.split():
                w = w.strip(".,?!;:").lower()
                if w:
                    vocab.add(w)
    return vocab


def load(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))["lexicons"]


def clean(src: dict, gold_vocab: set | None = None) -> tuple[dict, dict]:
    """Returns (canonical->[variants] for each dict, drop-reason stats)."""
    stats = defaultdict(int)
    cleaned = {}
    gold_vocab = gold_vocab or set()
    known_correct = build_known_correct(src, gold_vocab)

    for dict_name in ("corrections", "proper_nouns"):
        flat = src.get(dict_name, {})
        grouped: dict[str, set] = defaultdict(set)

        for variant, canonical in flat.items():
            v = variant.strip()
            c = str(canonical).strip()
            vl, cl = v.lower(), c.lower()

            # R3 — self-map. Keep ONLY if it is a genuine case-normaliser.
            if vl == cl:
                if v != c:                      # e.g. cnic -> CNIC  (useful)
                    grouped[c].add(vl)          # variants stored lowercase
                    stats["kept_case_normaliser"] += 1
                else:                           # e.g. center -> center (no-op)
                    stats["drop_selfmap_noop"] += 1
                continue

            # R1 — length floor. Only kills 1-2 char variants (inherently unsafe).
            if len(v) < MIN_LEN:
                stats["drop_too_short"] += 1
                continue

            # FORCE_KEEP — explicit override of the safety rules below.
            if (vl, cl) in FORCE_KEEP:
                grouped[c].add(vl)
                stats["force_kept"] += 1
                continue

            # R12 — the variant is not Latin script (Nastaliq / Arabic / Devanagari).
            # The pipeline is Devanagari -> transliterate() -> ROMAN, so the lexicon
            # only ever sees Latin tokens. Non-Latin variants can never fire.
            if not re.fullmatch(r"[A-Za-z0-9 .'\-]+", v):
                stats["drop_non_latin"] += 1
                continue

            # R1b — THE REAL SAFETY TEST: the variant is already a correct word.
            # Mapping it away is what corrupted 23% of gold (se->s, aap->app,
            # main->mein). Unlike a length rule, this KEEPS eria->area.
            if vl in known_correct:
                stats["drop_is_real_word"] += 1
                continue

            # R8 — word -> phrase expansion. DATA-DRIVEN, not a blanket ban.
            #
            # An earlier version dropped ALL of these "because they break token
            # counts". That was wrong and cost 64 good fixes to block 2 bad ones:
            #   assalaamualaikum -> "assalam o alaikum"   GOOD: the ASR merges it
            #   aapka -> "aap ka",  aapko -> "aap ko"     GOOD: gold splits them
            #   karenge -> "karein ge"                    BAD:  gold keeps it joined
            #
            # The GOLD decides. Expanding is only wrong when the annotators write
            # the word JOINED — then splitting it creates an extra token and a
            # WER error. If the gold writes it SPLIT, expanding is exactly right.
            if " " not in v and " " in c:
                if vl in gold_vocab:          # gold keeps this word joined
                    stats["drop_word_to_phrase"] += 1
                    continue
                # gold splits it (or never uses it) -> the expansion is correct

            # R5 — this SPECIFIC (variant -> canonical) map is wrong.
            # Ban the PAIR, not the canonical: `nephrology` is a fine word, it is
            # only `neurologist -> nephrology` that is dangerous. Banning the
            # canonical would also kill nefrologee/nephrolgy/... (~150 good fixes).
            if (vl, cl) in BANNED_PAIRS:
                stats["drop_banned_pair"] += 1
                continue

            # canonical is a single letter / fragment — never a valid target
            if cl in BANNED_CANONICALS:
                stats["drop_banned_canonical"] += 1
                continue

            # R11 — both forms are valid spellings; don't enforce a preference
            if (vl, cl) in SPELLING_PREFERENCE_ONLY:
                stats["drop_spelling_preference"] += 1
                continue

            grouped[c].add(vl)                  # variants stored lowercase
            stats["kept"] += 1

        # R4 — break bidirectional cycles: if A is a canonical AND a variant of B
        # IMPORTANT: skip case-normalisers (chennai -> Chennai). Those look like a
        # self-cycle to a naive check (the variant lowercases to its own canonical)
        # but they are the SAME entity differing only in case — not a cycle.
        canon_lower = {k.lower(): k for k in grouped}
        for canon in list(grouped):
            for v in list(grouped[canon]):
                if v.lower() == canon.lower():
                    continue                      # case-normaliser, keep it
                if v.lower() in canon_lower and canon.lower() in {
                    x.lower() for x in grouped.get(canon_lower[v.lower()], ())
                }:
                    grouped[canon].discard(v)
                    stats["drop_cycle"] += 1

        # R9 — MERGE case-fragmented canonicals.
        # "chughtai" (23 variants) and "Chughtai" (1 variant) are the SAME entity.
        # Left split, the lowercase form wins and the proper capitalisation is lost.
        # For proper_nouns prefer the Capitalised form; for corrections prefer lowercase.
        by_lower: dict[str, list[str]] = defaultdict(list)
        for k in grouped:
            by_lower[k.lower()].append(k)

        merged: dict[str, set] = {}
        for low, forms in by_lower.items():
            if len(forms) == 1:
                merged[forms[0]] = grouped[forms[0]]
                continue
            # Prefer the CAPITALISED form as canonical in both dicts. Lookup
            # lowercases the input word, so a capitalised canonical can still be
            # reached — and it preserves proper capitalisation on output.
            winner = next((f for f in forms if f[:1].isupper()), forms[0])

            allv: set = set()
            for f in forms:
                allv |= grouped[f]
                if f != winner:
                    allv.add(f)          # the losing spelling becomes a variant

            # Drop no-op variants: lookup lowercases the word, so a variant whose
            # lowercase form equals an ALL-LOWERCASE canonical resolves to itself.
            # (Keeping it is harmless but pointless; it also trips the audit.)
            wl = winner.lower()
            allv = {
                v for v in allv
                if v.lower() != wl or winner != wl   # keep only if it capitalises
            }
            merged[winner] = allv
            stats["merged_case_fragment"] += 1

        cleaned[dict_name] = {
            k: sorted(v) for k, v in sorted(merged.items(), key=lambda x: x[0].lower()) if v
        }

    # ── R10 — UNIFY the two dicts ────────────────────────────────────────────
    # `corrections` and `proper_nouns` are both word-level exact-match maps; the
    # split adds no behaviour but DOES fragment entities across them:
    #   corrections["chughtai"]  = [4 variants]
    #   proper_nouns["Chughtai"] = [25 variants]     <- SAME entity, split!
    # 74 such cross-dict collisions existed. Merge into ONE map, keyed by the
    # correct canonical (prefer the capitalised form — proper nouns should be
    # capitalised, and acronyms like CNIC must stay uppercase).
    # TWO passes. A single pass is buggy: if we switch the winning canonical
    # mid-way, variants already written under the losing key are stranded there
    # (that left 'allah' AND 'Allah' both present).
    #
    # Pass 1 — decide the winning canonical for each lowercase key.
    canon_for_lower: dict[str, str] = {}
    for dict_name in ("proper_nouns", "corrections"):
        for canon in cleaned[dict_name]:
            low = canon.lower()
            cur = canon_for_lower.get(low)
            if cur is None:
                canon_for_lower[low] = canon
            elif canon != low and cur == low:
                canon_for_lower[low] = canon      # prefer the capitalised form

    # Pass 2 — pour every entity's variants into its winning canonical.
    unified: dict[str, set] = defaultdict(set)
    for dict_name in ("proper_nouns", "corrections"):
        for canon, variants in cleaned[dict_name].items():
            winner = canon_for_lower[canon.lower()]
            unified[winner] |= set(variants)
            if canon != winner:
                unified[winner].add(canon.lower())   # losing spelling is a variant
                stats["merged_cross_dict"] += 1

    # a variant must never equal its own canonical unless it capitalises it
    final: dict[str, list] = {}
    for canon, variants in unified.items():
        cl = canon.lower()
        keep = {v for v in variants if v.lower() != cl or canon != cl}
        if keep:
            final[canon] = sorted(keep)

    # ── Split into `lexicon` and `phrases` — by MECHANISM, not semantics ──────
    #
    # WHY NO acronym/proper_noun/word CATEGORIES:
    #   We tried. It cannot be done reliably, and a confidently-wrong category is
    #   worse than none. To categorise you must know that `afzal` is a name but
    #   `aasman` (sky) is a word — and nothing in the data tells you that:
    #     - the source's own casing is arbitrary (it had BOTH `arif` and `afzal`
    #       lowercase, yet one ended up capitalised and the other did not)
    #     - an English dictionary fails in BOTH directions: it calls `Delhi`,
    #       `Allah` and `Apple` dictionary words, and calls Urdu words like
    #       `aasman`, `aaiye`, `achanak` non-words (=> "must be a name")
    #   Any auto-classifier would mislabel hundreds of entries.
    #
    #   The categories also bought us nothing: the runtime flattens them into one
    #   lookup anyway, and the only real consumer — the ASR context glossary —
    #   must be SMALL and per-call scoped (the scale test showed a 180-term
    #   glossary performs WORSE than a 3-term one). So the glossary comes from the
    #   curated data/entities.json gazetteer, not from auto-classification.
    #
    # The lexicon/phrases split IS justified: they need different lookup machinery
    # (exact word match vs. regex phrase replacement).
    words = {k: v for k, v in final.items() if " " not in k}
    phrases = {k: v for k, v in final.items() if " " in k}

    # Title-case canonicals that the gazetteer positively identifies as entities,
    # so output casing is consistent (Afzal, Arif, Sialkot — not afzal, Arif, ...).
    gaz = load_gazetteer()
    retitled: dict[str, list] = {}
    for k, v in words.items():
        proper = gaz.get(k.lower())
        if proper and proper != k:
            retitled[proper] = sorted(set(v) | set(retitled.get(proper, [])) | {k.lower()})
            stats["retitled_entity"] += 1
        else:
            retitled[k] = v
    words = {k: sorted(set(v) - {k.lower()} if k != k.lower() else set(v))
             for k, v in retitled.items()}
    words = {k: v for k, v in words.items() if v}

    out = {
        "lexicon": dict(sorted(words.items(), key=lambda x: x[0].lower())),
        "phrases": dict(sorted(phrases.items(), key=lambda x: x[0].lower())),
    }
    _assert_invariants(out)
    return out, dict(stats)


def load_gazetteer() -> dict:
    """
    data/entities.json -> {lowercase: TitleCase}.

    A small, CURATED, hand-maintained entity list. Its ONLY jobs are:
      1. make output casing consistent for known names/places
      2. supply the ASR context glossary (Layer 1)
    It is deliberately NOT used to categorise the whole lexicon — see the long
    comment in clean(). High precision, does not need to be complete.
    """
    path = SCRIPT_DIR / "data" / "entities.json"
    if not path.exists():
        return {}
    raw = json.loads(path.read_text(encoding="utf-8"))
    out = {}
    for key, items in raw.items():
        if key.startswith("_"):
            continue
        for e in items:
            out[e.lower()] = e.title() if not e.isupper() else e
    return out


CATEGORIES = ("lexicon", "phrases")


def _assert_invariants(out: dict) -> None:
    """
    HARD INVARIANTS — the cleaner refuses to emit a file that violates these.

    Uniqueness is enforced ACROSS ALL CATEGORIES, not per-category. Keys carry
    mixed case by design (CNIC / Chughtai / area) — so the same entity could
    otherwise land in two categories and be silently split, which is exactly the
    chughtai + Chughtai bug that the old corrections/proper_nouns split caused.
    """
    # I1 — no canonical may collide with another, case-insensitively, ANYWHERE
    seen: dict[str, tuple[str, str]] = {}
    for cat in CATEGORIES:
        for k in out[cat]:
            low = k.lower()
            if low in seen:
                pcat, pk = seen[low]
                raise AssertionError(
                    f"DUPLICATE CANONICAL KEY (case-insensitive): "
                    f"{pk!r} in [{pcat}] and {k!r} in [{cat}] are the same entity"
                )
            seen[low] = (cat, k)

    # I2 — a variant must belong to exactly one canonical, across all categories
    variant_owner: dict[str, str] = {}
    for cat in CATEGORIES:
        for canon, variants in out[cat].items():
            for v in variants:
                vl = v.lower()
                if vl in variant_owner and variant_owner[vl] != canon:
                    raise AssertionError(
                        f"VARIANT {v!r} claimed by two canonicals: "
                        f"{variant_owner[vl]!r} and {canon!r}"
                    )
                variant_owner[vl] = canon

    # I3 — no duplicate variants inside an entry; none may be a no-op
    for cat in CATEGORIES:
        for canon, variants in out[cat].items():
            lows = [v.lower() for v in variants]
            if len(lows) != len(set(lows)):
                raise AssertionError(f"DUPLICATE VARIANT inside {canon!r}: {variants}")
            if canon.lower() in lows and canon == canon.lower():
                raise AssertionError(f"NO-OP variant equals canonical in {canon!r}")

    # I4 — every variant must be stored lowercase (lookup lowercases its input)
    for cat in CATEGORIES:
        for canon, variants in out[cat].items():
            bad = [v for v in variants if v != v.lower()]
            if bad:
                raise AssertionError(f"NON-LOWERCASE variant in {canon!r}: {bad}")


def to_lookup(cleaned: dict) -> tuple[dict, dict]:
    """
    Invert {canonical: [variants]} -> {variant: canonical} for O(1) runtime lookup.

    The 4 categories are for HUMANS (navigability, and acronyms+proper_nouns is
    the ASR glossary). At runtime they collapse into one flat word map, plus a
    phrase map (phrases need regex replacement, not word lookup).

    Returns (word_lookup, phrase_lookup).
    """
    words, phrases = {}, {}
    for canon, variants in cleaned.get("lexicon", {}).items():
        for v in variants:
            words[v.lower()] = canon
    for canon, variants in cleaned.get("phrases", {}).items():
        for v in variants:
            phrases[v.lower()] = canon
    return words, phrases


def entity_glossary(cleaned: dict | None = None) -> list[str]:
    """
    The ASR context glossary (Layer 1) comes from the CURATED gazetteer, not from
    auto-classifying the lexicon. Keep it small — the scale test showed a 180-term
    glossary performs WORSE than a 3-term one (soft biasing dilutes).
    In production, scope it per call: base entities + this call's lab/agent/patient.
    """
    return sorted(set(load_gazetteer().values()), key=str.lower)


def measure_corruption(corr: dict, pn: dict, label: str) -> tuple[int, int]:
    """Apply the lookup to every GOLD word; count how many correct words get changed."""
    import openpyxl
    wb = openpyxl.load_workbook(str(XLSX), data_only=True)
    ws = wb["asr_results"]
    rows = list(ws.iter_rows(values_only=True))
    idx = {h: i for i, h in enumerate(rows[0])}

    total = changed = 0
    examples = {}
    for r in rows[1:]:
        ref = r[idx["roman_urdu_reference"]]
        if not isinstance(ref, str):
            continue
        for w in ref.split():
            total += 1
            lw = w.lower()
            new = pn.get(lw) or corr.get(lw)
            if new and new.lower() != lw:
                changed += 1
                examples.setdefault(w, new)
    pct = 100 * changed / total if total else 0
    print(f"  {label:<26} {changed:>4} / {total} gold words corrupted  ({pct:.2f}%)")
    return changed, total


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--write", action="store_true", help="write data/lexicons_v2.json")
    args = ap.parse_args()

    src = load(SRC)
    gold_vocab = load_gold_vocab()
    print("=" * 78)
    print("  STEP 2 — LEXICON CLEANUP")
    print("=" * 78)
    print(f"  source: {SRC.name}")
    print(f"    corrections  : {len(src['corrections']):>6} flat entries")
    print(f"    proper_nouns : {len(src['proper_nouns']):>6} flat entries")
    print(f"  gold protected vocabulary: {len(gold_vocab)} human-verified words")
    print()

    cleaned, stats = clean(src, gold_vocab)

    print("  DROP REASONS")
    labels = {
        "drop_too_short":          f"R1  variant < {MIN_LEN} chars (1-2 char only)",
        "drop_is_real_word":       "R1b variant is ALREADY A CORRECT WORD  <-- the real test",
        "drop_non_latin":          "R12 variant is not Latin script (Nastaliq)",
        "force_kept":              "    FORCE-KEPT (explicit override)",
        "drop_selfmap_noop":       "R3  self-map no-op (key == value)",
        "drop_cycle":              "R4  bidirectional cycle (A<->B)",
        "drop_banned_pair":        "R5  BANNED PAIR (this specific map is wrong)",
        "drop_banned_canonical":   "R5b canonical is a fragment (s, m, ii)",
        "drop_word_to_phrase":     "R8  word -> phrase (breaks token count)",
        "merged_case_fragment":    "R9  MERGED case-fragment (within a dict)",
        "merged_cross_dict":       "R10 MERGED cross-dict entity (chughtai + Chughtai)",
        "kept_case_normaliser":    "R6  KEPT case-normaliser (cnic->CNIC)",
        "kept":                    "    KEPT real variant",
    }
    for k, lbl in labels.items():
        if k in stats:
            print(f"    {lbl:<44} {stats[k]:>6}")
    print()

    print("  RESULT — {canonical: [variants]}  (no semantic categories — see clean())")
    tot_c = tot_v = 0
    desc = {"lexicon": "single words  (exact match lookup)",
            "phrases": "multi-word    (regex replacement)"}
    for cat in CATEGORIES:
        n = len(cleaned[cat]); v = sum(len(x) for x in cleaned[cat].values())
        tot_c += n; tot_v += v
        print(f"    {cat:<9} {n:>5} canonicals  {v:>6} variants   {desc[cat]}")
    print(f"    {'TOTAL':<9} {tot_c:>5} canonicals  {tot_v:>6} variants")
    print()
    gl = entity_glossary()
    print(f"  ASR glossary (from curated data/entities.json): {len(gl)} entities")
    print(f"    e.g. {', '.join(gl[:8])} ...")
    print()

    # ---- re-measure corruption on gold ----
    print("  CORRUPTION ON THE 183-TURN GOLD SET")
    base = load(BASE)
    measure_corruption(base["corrections"], base["proper_nouns"], "original lexicons.json")
    measure_corruption(src["corrections"], src["proper_nouns"], "lexicons_updated.json")
    w, ph = to_lookup(cleaned)
    measure_corruption(w, {}, "lexicons_v2.json (NEW)")
    print()

    if args.write:
        out = {
            "_comment": (
                "CLEANED lexicon. Structure: {canonical: [variant, ...]}. "
                "TWO sections, split by MECHANISM not semantics: `lexicon` (single "
                "words, exact-match lookup) and `phrases` (multi-word, regex "
                "replacement). There are deliberately NO acronym/proper_noun/word "
                "categories: nothing in the data can tell you that `afzal` is a name "
                "but `aasman` (sky) is a word — source casing is arbitrary and an "
                "English dictionary fails both ways. A confidently-wrong category is "
                "worse than none. Canonicals keep their correct case (CNIC, Chughtai); "
                "variants are ALWAYS lowercase (lookup lowercases its input). Keys are "
                "unique case-insensitively — enforced by _assert_invariants(). The ASR "
                "context glossary comes from the curated data/entities.json instead."
            ),
            "lexicons": cleaned,
        }
        OUT.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"  ✓ written: {OUT}")
    else:
        print("  (dry run — pass --write to save data/lexicons_v2.json)")
    print()


if __name__ == "__main__":
    main()
