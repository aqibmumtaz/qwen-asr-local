#!/usr/bin/env python3
"""
Generate an annotation-ready XLSX from ASR confidence JSON.

Reads benchmark_results/cll_confidence_raw.json (output of hf_asr_with_confidence)
and produces an Excel workbook with:
  - Sheet "Stats": summary statistics + confidence distribution
  - Sheet "Words": per-word table sorted by confidence (worst first), with empty
    correction columns for human annotators

Usage:
    python3 generate_annotation_sheet.py [input.json] [output.xlsx]
"""

import json
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    import subprocess

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# Transliteration imports
sys.path.insert(0, str(Path(__file__).resolve().parent))
from hindi_to_roman_urdu import (
    transliterate as to_roman_urdu,
    _transliterate_raw,
    _normalize_endings,
)
from indo_arabic_transliteration.hindustani import HindustaniTransliterator

_nastaliq = HindustaniTransliterator()

LOW_CONF_THRESHOLD = 0.65


def _raw_roman(hindi: str) -> str:
    return _normalize_endings(_transliterate_raw(hindi))


def generate_sheet(input_json: Path, output_xlsx: Path):
    data = json.loads(input_json.read_text())
    words = data["words"]
    elapsed = data["elapsed"]
    audio_file = Path(data["audio_file"]).name
    full_text = data["text"]

    # Sort by min_conf ascending (worst first for annotation)
    words_sorted = sorted(words, key=lambda w: w["min_conf"])

    # Stats
    total_words = len(words)
    low_words = sum(1 for w in words if w["min_conf"] <= LOW_CONF_THRESHOLD)
    pct_low = (low_words / total_words * 100) if total_words else 0

    # Confidence buckets
    buckets = {"0.0–0.50": 0, "0.50–0.65": 0, "0.65–0.80": 0, "0.80–1.00": 0}
    for w in words:
        c = w["min_conf"]
        if c <= 0.50:
            buckets["0.0–0.50"] += 1
        elif c <= 0.65:
            buckets["0.50–0.65"] += 1
        elif c <= 0.80:
            buckets["0.65–0.80"] += 1
        else:
            buckets["0.80–1.00"] += 1

    # Create workbook
    wb = openpyxl.Workbook()

    # ── Stats Sheet ──────────────────────────────────────────────────────
    ws_stats = wb.active
    ws_stats.title = "Stats"

    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    low_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    ws_stats["A1"] = "ASR Confidence Analysis"
    ws_stats["A1"].font = title_font

    stats_data = [
        ("Audio File", audio_file),
        ("Full Hindi Text", full_text),
        ("Inference Time", f"{elapsed:.1f}s"),
        ("Total Words", total_words),
        ("LOW Confidence Words (≤0.65)", low_words),
        ("% Flagged", f"{pct_low:.1f}%"),
        ("", ""),
        ("Confidence Distribution", ""),
    ]
    for i, (label, value) in enumerate(stats_data, start=3):
        ws_stats[f"A{i}"] = label
        ws_stats[f"A{i}"].font = Font(bold=True)
        ws_stats[f"B{i}"] = value

    row = 3 + len(stats_data)
    for bucket, count in buckets.items():
        ws_stats[f"A{row}"] = f"  {bucket}"
        ws_stats[f"B{row}"] = count
        ws_stats[f"C{row}"] = f"({count/total_words*100:.0f}%)" if total_words else ""
        row += 1

    ws_stats.column_dimensions["A"].width = 30
    ws_stats.column_dimensions["B"].width = 80

    # ── Words Sheet (annotation-ready) ───────────────────────────────────
    ws_words = wb.create_sheet("Annotation")

    headers = [
        "#",
        "Hindi (ASR)",
        "Raw Roman",
        "Roman Urdu",
        "Nastaliq",
        "Min Conf",
        "Geo Conf",
        "Tokens",
        "Flag",
        "Corrected Roman Urdu",
        "Corrected Nastaliq",
        "Notes",
    ]

    # Write headers
    for col, h in enumerate(headers, 1):
        cell = ws_words.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Write word rows (sorted by confidence — worst first)
    for i, w in enumerate(words_sorted, start=1):
        hindi = w["text"]
        flag = "LOW" if w["min_conf"] <= LOW_CONF_THRESHOLD else ""
        row_data = [
            i,
            hindi,
            _raw_roman(hindi),
            to_roman_urdu(hindi),
            _nastaliq.transliterate_from_hindi_to_urdu(hindi),
            round(w["min_conf"], 3),
            round(w["geo_conf"], 3),
            w["n_tokens"],
            flag,
            "",  # Corrected Roman Urdu (empty for annotator)
            "",  # Corrected Nastaliq (empty for annotator)
            "",  # Notes (empty for annotator)
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws_words.cell(row=i + 1, column=col, value=val)
            if flag == "LOW":
                cell.fill = low_fill

    # Column widths
    col_widths = [4, 18, 16, 16, 16, 9, 9, 7, 5, 22, 22, 20]
    for i, w in enumerate(col_widths, 1):
        ws_words.column_dimensions[get_column_letter(i)].width = w

    # Freeze header row
    ws_words.freeze_panes = "A2"

    # ── Original Order Sheet (for reference) ─────────────────────────────
    ws_orig = wb.create_sheet("Original Order")

    for col, h in enumerate(headers, 1):
        cell = ws_orig.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for i, w in enumerate(words, start=1):
        hindi = w["text"]
        flag = "LOW" if w["min_conf"] <= LOW_CONF_THRESHOLD else ""
        row_data = [
            i,
            hindi,
            _raw_roman(hindi),
            to_roman_urdu(hindi),
            _nastaliq.transliterate_from_hindi_to_urdu(hindi),
            round(w["min_conf"], 3),
            round(w["geo_conf"], 3),
            w["n_tokens"],
            flag,
            "",
            "",
            "",
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws_orig.cell(row=i + 1, column=col, value=val)
            if flag == "LOW":
                cell.fill = low_fill

    for i, w in enumerate(col_widths, 1):
        ws_orig.column_dimensions[get_column_letter(i)].width = w
    ws_orig.freeze_panes = "A2"

    # Save
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_xlsx))
    print(f"\n✓ Saved annotation sheet: {output_xlsx}")
    print(f"  - Stats sheet: summary + confidence distribution")
    print(
        f"  - Annotation sheet: {total_words} words sorted by confidence (worst first)"
    )
    print(f"    → {low_words} LOW words highlighted in red, ready for correction")
    print(f"  - Original Order sheet: words in transcription order")
    print(
        f"\nWorkflow: Open in Excel → Filter 'Flag' = LOW → Fill 'Corrected Roman Urdu' column"
    )


if __name__ == "__main__":
    input_json = (
        Path(sys.argv[1])
        if len(sys.argv) > 1
        else Path("benchmark_results/cll_confidence_raw.json")
    )
    output_xlsx = (
        Path(sys.argv[2])
        if len(sys.argv) > 2
        else Path("benchmark_results/cll_annotation_sheet.xlsx")
    )
    generate_sheet(input_json, output_xlsx)
