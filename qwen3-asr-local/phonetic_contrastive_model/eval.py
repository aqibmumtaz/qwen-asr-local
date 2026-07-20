"""
Evaluate the Phonetic Contrastive Model — the four numbers that decide it.

  python -m phonetic_contrastive_model.eval
  python -m phonetic_contrastive_model.eval --threshold 0.55

1. HELD-OUT variant recall   — unseen spellings of KNOWN canonicals -> right canonical?
                               (top-1 with abstain, top-1 raw, top-3). The generalisation number.
2. ABSTAIN safety            — real gold words that are NOT canonicals: % left UNCHANGED.
                               (the anti-corruption number.)
3. EXACT-NAME held-out       — held-out variants whose canonical is an entity name:
                               exact top-1 match. The decisive Siddique/Siddiqui test.
4. 80-call diff_words        — standalone corrector on top of v2 (resolver off) vs baseline.
                               The production-metric number.
"""
from __future__ import annotations

import argparse
import importlib
import os
import sys
from collections import Counter, defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT / "testing"))

import openpyxl  # noqa: E402

from phonetic_contrastive_model.corrector import PhoneticContrastiveCorrector  # noqa: E402
from phonetic_contrastive_model.data import make_splits  # noqa: E402
from test_accuracy import diff_words  # noqa: E402

XLSX = ROOT / "testing" / "lab_test_80_calls_urdu_roman_urdu.xlsx"


def gold_noncanon_words(canon_set):
    """Correct words humans wrote (benchmark) that are NOT canonicals -> abstain test."""
    wb = openpyxl.load_workbook(str(XLSX))
    ws = wb["Sheet1"]; rows = list(ws.iter_rows(values_only=True))
    idx = {h: i for i, h in enumerate(rows[0])}
    out = set()
    for r in rows[1:]:
        ref = r[idx["benchmark_roman_urdu"]]
        if isinstance(ref, str):
            for w in ref.split():
                w = w.strip(".,?!;:").lower()
                if w.isalpha() and len(w) >= 3 and w not in canon_set:
                    out.add(w)
    return sorted(out)


def eighty_call(corr):
    """diff_words: v2 (resolver off) alone vs v2 + contrastive corrector on top."""
    os.environ["LEXICON"] = "v2"; os.environ["RESOLVER"] = "0"
    import hindi_to_roman_urdu as H
    importlib.reload(H)
    wb = openpyxl.load_workbook(str(XLSX)); ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True)); idx = {h: i for i, h in enumerate(rows[0])}
    calls = defaultdict(list)
    for r in rows[1:]:
        calls[r[idx["call_id"]]].append(r)
    for cid in calls:
        calls[cid].sort(key=lambda x: x[idx["chunk_index"]] or 0)
    base_m = base_t = corr_m = corr_t = 0
    for cid, cr in calls.items():
        bench = next((x[idx["benchmark_roman_urdu"]] for x in cr
                      if isinstance(x[idx["benchmark_roman_urdu"]], str)
                      and x[idx["benchmark_roman_urdu"]].strip()), None)
        if not bench:
            continue
        hindi = " ".join(str(x[idx["model_output_hindi"]]).strip() for x in cr
                         if isinstance(x[idx["model_output_hindi"]], str)
                         and x[idx["model_output_hindi"]].strip())
        base = H.transliterate(hindi)
        fixed = corr.resolve_text(base)
        b = diff_words(bench, base); base_m += b.matched; base_t += b.total
        c = diff_words(bench, fixed); corr_m += c.matched; corr_t += c.total
    return 100 * base_m / base_t, 100 * corr_m / corr_t


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--threshold", type=float, default=0.60)
    ap.add_argument("--seed", type=int, default=13)
    args = ap.parse_args()

    sp = make_splits(seed=args.seed)
    heldout = sp["heldout_pairs"]
    entity_canon = sp["entity_canonicals"]
    corr = PhoneticContrastiveCorrector.load(threshold=args.threshold)

    # 1 + 3 — held-out recall (all, and names only) --------------------------
    top1 = top1_raw = top3 = 0
    n_names = names_top1 = 0
    for variant, canon in heldout:
        out = corr.resolve_word(variant)
        tk = [c for c, _ in corr.topk(variant, 3)]
        raw1 = tk[0] if tk else ""
        if out.lower() == canon.lower():
            top1 += 1
        if raw1.lower() == canon.lower():
            top1_raw += 1
        if canon.lower() in [t.lower() for t in tk]:
            top3 += 1
        if canon.lower() in entity_canon:
            n_names += 1
            if out.lower() == canon.lower():
                names_top1 += 1
    n = len(heldout)
    n_common = n - n_names
    common_top1 = top1 - names_top1

    # 2 — abstain safety -----------------------------------------------------
    canon_set = {c.lower() for c in sp["canonicals"]}
    words = gold_noncanon_words(canon_set)
    unchanged = sum(1 for w in words if corr.resolve_word(w).lower() == w)
    changed = [(w, corr.resolve_word(w)) for w in words if corr.resolve_word(w).lower() != w]

    # 4 — 80-call diff_words -------------------------------------------------
    base_acc, corr_acc = eighty_call(corr)

    print("=" * 74)
    print(f"  PHONETIC CONTRASTIVE MODEL — eval  (threshold={args.threshold})")
    print("=" * 74)
    print(f"\n  1. HELD-OUT variant recall  (n={n} unseen spellings)")
    print(f"     top-1 (with abstain) : {100*top1/n:5.1f}%")
    print(f"     top-1 (raw nearest)  : {100*top1_raw/n:5.1f}%")
    print(f"     top-3                : {100*top3/n:5.1f}%")
    print(f"\n  3. by type:")
    print(f"     COMMON words  top-1  : {100*common_top1/max(n_common,1):5.1f}%  (n={n_common})")
    print(f"     NAMES         top-1  : {100*names_top1/max(n_names,1):5.1f}%  (n={n_names})   <- decisive")
    print(f"\n  2. ABSTAIN safety  (real gold words, NOT canonicals, n={len(words)})")
    print(f"     left UNCHANGED       : {100*unchanged/len(words):5.1f}%")
    print(f"     wrongly changed      : {len(changed)}   e.g. {changed[:6]}")
    print(f"\n  4. 80-call diff_words (standalone on top of v2, resolver off)")
    print(f"     v2 baseline          : {base_acc:5.2f}%")
    print(f"     v2 + contrastive     : {corr_acc:5.2f}%   ({corr_acc-base_acc:+.2f})")
    print()
    print(f"  corrector stats: {dict(corr.stats)}")
    print()


if __name__ == "__main__":
    main()
