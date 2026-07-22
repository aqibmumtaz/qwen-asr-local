#!/usr/bin/env python3
"""
Benchmark on the 80-call lab-test set, CALL-LEVEL, on the model_output_hindi column.

Ground truth is per-call (benchmark_roman_urdu, on chunk_index==0). model_output_*
are per-chunk. So for each call we transliterate each chunk's model_output_hindi,
join the per-chunk Roman-Urdu in chunk order, and score that against the call's
benchmark. The per-chunk Roman-Urdu we produce is ALSO written out verbatim, so the
"model_output" sheet column is exactly the text that was scored.

Metric: the dev's testing/test_accuracy.py :: diff_words   (fuzzy word RECALL, >=0.70)
Also reported: classic edit-distance WER accuracy (1 - edit/ref_words), same tokeniser.

Configs:
  prev     previous model's own Roman  (model_output_roman_urdu, no transliterate)
  v2       transliterate() + v2 exact lexicon, no phonetic          (text baseline)
  v22ph    transliterate() + v2.2 lexicon + phonetic model @0.90    (PRODUCTION, .env)

Output workbook (a COPY; the original is untouched):
  - Sheet1                original sheet, unchanged
  - model_v22_phonetic    NEW sheet: per-chunk output of the PRODUCTION model
                          (v2.2 + phonetic) next to model_output_hindi
  - benchmark_summary     NEW sheet: per-call scores for every config

  python benchmark/benchmark_baseline.py
"""
import importlib
import os
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
sys.path.insert(0, str(HERE))   # test_accuracy.py
sys.path.insert(0, str(ROOT))   # hindi_to_roman_urdu.py

from test_accuracy import diff_words, normalize_tokens

XLSX = HERE / "lab_test_80_calls_urdu_roman_urdu.xlsx"
OUT = HERE / "lab_test_80_calls_urdu_roman_urdu_benchmarked.xlsx"

# The production config, mirrors .env. Change here if .env changes.
PROD = {"LEXICON": "v22", "PHONETIC": "1", "PHONETIC_THRESHOLD": "0.9", "RESOLVER": "0"}


def wer_acc(benchmark: str, hypothesis: str) -> tuple[int, int]:
    """Classic word edit-distance. Returns (edits, ref_len) using the dev tokeniser."""
    ref = normalize_tokens(benchmark)
    hyp = normalize_tokens(hypothesis)
    if not ref:
        return (0, 0)
    dp = list(range(len(hyp) + 1))
    for rw in ref:
        nd = [dp[0] + 1] + [0] * len(hyp)
        for j, hw in enumerate(hyp):
            nd[j + 1] = min(dp[j] + (0 if hw == rw else 1), dp[j + 1] + 1, nd[j] + 1)
        dp = nd
    return (dp[len(hyp)], len(ref))


def load_calls():
    wb = openpyxl.load_workbook(str(XLSX))
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    idx = {h: i for i, h in enumerate(rows[0])}
    calls = defaultdict(list)
    for r in rows[1:]:
        calls[r[idx["call_id"]]].append(r)
    for cid in calls:
        calls[cid].sort(key=lambda r: (r[idx["chunk_index"]] if r[idx["chunk_index"]] is not None else 0))
    return wb, ws, idx, calls, rows


def set_env(cfg: dict):
    for k in ("LEXICON", "PHONETIC", "PHONETIC_THRESHOLD", "RESOLVER"):
        os.environ.pop(k, None)
    os.environ.update(cfg)


def transliterate_for(cfg: dict):
    """Set env + reload the pipeline so its module globals (lexicon, phonetic model,
    resolver flag) are rebuilt for THIS config, then return its transliterate()."""
    set_env(cfg)
    import hindi_to_roman_urdu as H
    importlib.reload(H)
    return H.transliterate


def main():
    wb, ws, idx, calls, rows = load_calls()

    # ONE row-object per data row of Sheet1, in the ORIGINAL row order — so the output
    # sheet lines up one-to-one with Sheet1 (every model_output_hindi row, benchmarked
    # or not, gets its corresponding output row).
    all_chunks = []
    for r in rows[1:]:
        h = r[idx["model_output_hindi"]]
        pv = r[idx["model_output_roman_urdu"]]
        au = r[idx["actual_urdu_transcript"]]
        br = r[idx["benchmark_roman_urdu"]]
        all_chunks.append({
            "cid": r[idx["call_id"]],
            "ci": r[idx["chunk_index"]],
            "audio": r[idx["audio_name"]],
            "hindi": h.strip() if isinstance(h, str) else "",
            "prev": pv.strip() if isinstance(pv, str) else "",
            "actual_urdu": au.strip() if isinstance(au, str) else "",
            "bench_roman": br.strip() if isinstance(br, str) else "",
        })

    # Run each transliterate config in its OWN pass. transliterate() reads module
    # globals at call time and importlib.reload mutates the shared module dict in
    # place, so compute ALL rows for one config before moving to the next.
    CONFIGS = [
        ("v2",    {"LEXICON": "v2",  "PHONETIC": "",  "RESOLVER": "0"}),
        ("v22ph", PROD),
    ]
    for key, cfg in CONFIGS:
        tr = transliterate_for(cfg)
        for ch in all_chunks:
            ch[key] = tr(ch["hindi"]) if ch["hindi"] else ""

    # group by call for scoring; only calls WITH a benchmark are scored.
    KEYS = ["prev", "v2", "v22ph"]
    by_call = defaultdict(list)
    for ch in all_chunks:
        by_call[ch["cid"]].append(ch)
    per_call = {}
    for cid, crows in calls.items():
        bench = next((r[idx["benchmark_roman_urdu"]] for r in crows
                      if isinstance(r[idx["benchmark_roman_urdu"]], str)
                      and r[idx["benchmark_roman_urdu"]].strip()), None)
        if not bench:
            continue
        chunks = sorted(by_call[cid], key=lambda c: c["ci"] if c["ci"] is not None else 0)
        d = {"bench": bench}
        for k in KEYS:
            d[k] = " ".join(ch[k] for ch in chunks if ch.get(k)).strip()
        d["acc"] = {k: diff_words(d["bench"], d[k]) for k in KEYS}
        d["wer"] = {k: wer_acc(d["bench"], d[k]) for k in KEYS}
        per_call[cid] = d

    # ---- aggregate ----
    def agg(cfg):
        m = sum(d["acc"][cfg].matched for d in per_call.values())
        t = sum(d["acc"][cfg].total for d in per_call.values())
        mean = sum(d["acc"][cfg].accuracy for d in per_call.values()) / len(per_call)
        we = sum(d["wer"][cfg][0] for d in per_call.values())
        wt = sum(d["wer"][cfg][1] for d in per_call.values())
        return dict(corpus=100 * m / t, mean=mean, matched=m, total=t,
                    wer=100 * max(0.0, 1 - we / wt))

    A = {c: agg(c) for c in KEYS}
    labels = {"prev":  "prev    (vendor roman, no fix)",
              "v2":    "v2      (exact lexicon only)",
              "v22ph": "v22ph   (v2.2 + phonetic @0.90) *"}

    print("=" * 80)
    print(f"  BENCHMARK — 80-call set, call-level, on model_output_hindi   ({len(per_call)} calls)")
    print("=" * 80)
    print(f"  metric: diff_words (fuzzy word recall >=0.70)  +  edit-distance WER acc")
    print(f"  * = production config (.env: {PROD})")
    print()
    print(f"  {'config':<34} {'diff_words':>11} {'mean/call':>10} {'WER acc':>9} {'vs prev':>8}")
    print(f"  {'-'*34} {'-'*11} {'-'*10} {'-'*9} {'-'*8}")
    for c in KEYS:
        d = A[c]
        vs = d["corpus"] - A["prev"]["corpus"]
        print(f"  {labels[c]:<34} {d['corpus']:>10.2f}% {d['mean']:>9.2f}% "
              f"{d['wer']:>8.2f}% {vs:>+7.2f}")
    print()
    print(f"  production gain (v22ph - v2): {A['v22ph']['corpus']-A['v2']['corpus']:+.2f} "
          f"pts diff_words, {A['v22ph']['wer']-A['v2']['wer']:+.2f} pts WER")
    print()

    # ---- write copy: SHEET 2 = per-chunk production model output ----
    if "model_v22_phonetic" in wb.sheetnames:
        del wb["model_v22_phonetic"]
    s2 = wb.create_sheet("model_v22_phonetic")
    s2.append(["audio_name", "call_id", "chunk_index",
               "actual_urdu_transcript", "benchmark_roman_urdu",
               "model_output_hindi", "model_output_roman_urdu",
               "model_output_v22_phonetic"])
    # one row per Sheet1 data row, SAME order — aligns 1:1 with model_output_hindi
    for ch in all_chunks:
        s2.append([ch["audio"], ch["cid"], ch["ci"],
                   ch["actual_urdu"], ch["bench_roman"],
                   ch["hindi"], ch["prev"], ch.get("v22ph", "")])

    # ---- SHEET 3 = per-call summary for every config ----
    if "benchmark_summary" in wb.sheetnames:
        del wb["benchmark_summary"]
    s3 = wb.create_sheet("benchmark_summary")
    s3.append(["call_id", "benchmark_roman_urdu",
               "out_prev", "out_v2", "out_v22ph",
               "acc_prev", "acc_v2", "acc_v22ph",
               "wer_prev", "wer_v2", "wer_v22ph"])

    def werp(e, n):
        return round(100 * max(0.0, 1 - e / n), 2) if n else None

    for cid, d in per_call.items():
        s3.append([cid, d["bench"], d["prev"], d["v2"], d["v22ph"],
                   round(d["acc"]["prev"].accuracy, 4),
                   round(d["acc"]["v2"].accuracy, 4),
                   round(d["acc"]["v22ph"].accuracy, 4),
                   werp(*d["wer"]["prev"]), werp(*d["wer"]["v2"]), werp(*d["wer"]["v22ph"])])
    # corpus totals row
    s3.append([])
    s3.append(["CORPUS (diff_words %)", "", round(A["prev"]["corpus"], 2),
               round(A["v2"]["corpus"], 2), round(A["v22ph"]["corpus"], 2),
               "", "", "", "", "", ""])
    s3.append(["CORPUS (WER acc %)", "", round(A["prev"]["wer"], 2),
               round(A["v2"]["wer"], 2), round(A["v22ph"]["wer"], 2),
               "", "", "", "", "", ""])

    wb.save(str(OUT))
    print(f"  written: {OUT.name}")
    print(f"    - Sheet1               (original, unchanged)")
    print(f"    - model_v22_phonetic   (per-chunk production output beside model_output_hindi)")
    print(f"    - benchmark_summary    (per-call scores, all configs)")
    print()


if __name__ == "__main__":
    main()
