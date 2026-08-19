#!/usr/bin/env python3
"""
Benchmark silence-aware ("dynamic") re-chunked audio against the gold benchmark,
using the HF remote /chughtai/v1/realtime endpoint (domain-tuned for this exact
Chughtai Lab call-center dataset -- see acoustic_contextual_biasing/asr.py).

Audio: benchmark/lab_test_80_audios_chunks_dynamic/<call_id>/chunk_*.wav
       (re-cut at RMS-silence minima near each ~25s mark, instead of a hard
       25.00s clock cut -- see rechunk_all_manifest.json for per-call shifts)

The /chughtai endpoint returns raw_text (Hindi) and romanized_text (Roman Urdu)
directly in each completed-transcription event -- used AS-IS here, no local
transliteration step for those two columns:
  model_output_hindi       <- server's raw_text
  model_output_roman_urdu  <- server's romanized_text
  model_output_v22_phonetic <- local v2.2 lexicon + phonetic (@0.90) correction
                                applied on top of model_output_hindi

No local ASR is used anywhere in this script -- HF remote only.

Processes ONE CALL AT A TIME: transcribe its chunks, build its rows, append to
both sheets, and SAVE the workbook immediately -- so progress is visible call
by call in the actual xlsx file, not just in terminal output.

Resumable: cached per-chunk to transcription_cache_dynamic_hf_chughtai.jsonl.

  python benchmark/benchmark_chunks_dynamic.py resume
  python benchmark/benchmark_chunks_dynamic.py resume --calls 5
  python benchmark/benchmark_chunks_dynamic.py redo
"""
from __future__ import annotations

import argparse
import atexit
import importlib
import json
import os
import signal
import sys
import time
import warnings
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

warnings.filterwarnings("ignore")
HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
sys.path.insert(0, str(ROOT)); sys.path.insert(0, str(HERE))

import openpyxl
from test_accuracy import diff_words, normalize_tokens

XLSX = HERE / "lab_test_80_calls_urdu_roman_urdu.xlsx"
OUT = HERE / "lab_test_80_calls_urdu_roman_urdu_benchmarked.xlsx"
AUDIO = HERE / "lab_test_80_audios_chunks_dynamic"
PIDFILE = HERE / ".benchmark_chunks_dynamic.pid"
VARIANT = "chughtai"   # domain-tuned endpoint: /chughtai/v1/audio/transcriptions

# HTTP endpoints -- both serve the same /chughtai/v1/audio/transcriptions route,
# verified to score comparably (67.17% HF vs 67.92% GPU on a spot-check call).
ENDPOINT_BASE = {
    "hf": "https://ebitlogix-qwen-asr-vlm-async-test.hf.space",
    "gpu": "http://192.168.99.117:7873",
}

SHEET_NAME = "model_v22_phonetic_chunks_dynamic"
SUMMARY_NAME = "benchmark_summary_chunks_dynamic"


def cache_path_for(endpoint: str) -> Path:
    return HERE / f"transcription_cache_dynamic_{endpoint}_chughtai_http.jsonl"


def _kill_previous():
    if PIDFILE.exists():
        try:
            old_pid = int(PIDFILE.read_text().strip())
            if old_pid != os.getpid():
                os.kill(old_pid, signal.SIGTERM)
                print(f"Killed previous run (PID {old_pid})", flush=True)
        except (ProcessLookupError, ValueError, PermissionError):
            pass
    PIDFILE.write_text(str(os.getpid()))


def _cleanup_pid():
    try:
        if PIDFILE.exists() and PIDFILE.read_text().strip() == str(os.getpid()):
            PIDFILE.unlink()
    except Exception:
        pass


# ── data loading ─────────────────────────────────────────────────────────────

def load_calls():
    wb = openpyxl.load_workbook(str(XLSX))
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    idx = {h: i for i, h in enumerate(rows[0])}
    calls = defaultdict(list)
    for r in rows[1:]:
        calls[r[idx["call_id"]]].append(r)
    return idx, calls


# ── cache ────────────────────────────────────────────────────────────────────

def read_cache(cache_file: Path) -> dict:
    out = {}
    if cache_file.exists():
        for line in cache_file.read_text(encoding="utf-8").splitlines():
            if line.strip():
                r = json.loads(line)
                out[r["key"]] = r
    return out


def append_cache(entry: dict, cache_file: Path):
    with cache_file.open("a", encoding="utf-8") as f:
        f.write(json.dumps(entry, ensure_ascii=False) + "\n")


# ── wer helper (same tokeniser as benchmark_baseline.py) ────────────────────

def wer_acc(benchmark: str, hypothesis: str) -> tuple[int, int]:
    ref = normalize_tokens(benchmark)
    hyp = normalize_tokens(hypothesis)
    if not ref:
        return (0, 0)
    dp = list(range(len(hyp) + 1))
    for rw in ref:
        nd = [dp[0] + 1] + [0] * len(hyp)
        for j, hw in enumerate(hyp):
            nd[j + 1] = min(dp[j] + (0 if hw == rw else 1), dp[j + 1] + 1, nd[j] + 1)
        dp = nd
    return (dp[len(hyp)], len(ref))


def werp(e, n):
    return round(100 * max(0.0, 1 - e / n), 2) if n else None


# ── main run: transcribe + write, ONE CALL AT A TIME ────────────────────────

def run(args):
    from acoustic_contextual_biasing.asr import RemoteHttpASR

    base = ENDPOINT_BASE[args.endpoint]
    cache_file = cache_path_for(args.endpoint)

    if args.restart and cache_file.exists():
        cache_file.unlink()
        print("Cache cleared (redo mode)", flush=True)

    idx, calls = load_calls()
    cache = read_cache(cache_file)
    call_dirs = {d.name: d for d in AUDIO.iterdir() if d.is_dir()}

    # local text post-processing only (no ASR): v2.2 lexicon + phonetic @0.90,
    # loaded once up front -- this is NOT running any model/ASR locally, it's
    # deterministic lexicon/phonetic correction on the server's own Hindi text.
    os.environ["LEXICON"] = "v22"; os.environ["RESOLVER"] = "0"
    os.environ["PHONETIC"] = "1"; os.environ["PHONETIC_THRESHOLD"] = "0.9"
    import hindi_to_roman_urdu as H
    importlib.reload(H)

    def make_client():
        return RemoteHttpASR(variant=VARIANT, base=base, timeout=60.0, retries=1, verbose=False)

    ordered_cids = [cid for cid in calls if str(cid) in call_dirs]
    if args.calls > 0:
        ordered_cids = ordered_cids[:args.calls]

    print(f"Endpoint: {args.endpoint} -> {base}/{VARIANT}/v1/audio/transcriptions", flush=True)
    print(f"Calls to process: {len(ordered_cids)}  workers(per-call)={args.workers}", flush=True)

    print(f"Loading workbook ({OUT.name}) once...", flush=True)
    wb = openpyxl.load_workbook(str(OUT))
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    s = wb.create_sheet(SHEET_NAME)
    headers = ["audio_name", "call_id", "chunk_index", "actual_urdu_transcript",
               "benchmark_roman_urdu", "model_output_hindi",
               "model_output_roman_urdu", "model_output_v22_phonetic"]
    s.append(headers)

    if SUMMARY_NAME in wb.sheetnames:
        del wb[SUMMARY_NAME]
    s2 = wb.create_sheet(SUMMARY_NAME)
    s2.append(["call_id", "benchmark_roman_urdu",
               "out_server_roman", "acc_server_roman", "wer_server_roman",
               "out_v22_phonetic", "acc_v22_phonetic", "wer_v22_phonetic"])
    wb.save(str(OUT))
    print(f"  sheets created, initial save done.", flush=True)

    agg = {"server_roman": [0, 0], "v22ph": [0, 0]}
    n_scored = 0

    for ci, cid in enumerate(ordered_cids, 1):
        cid_str = str(cid)
        cd = call_dirs[cid_str]
        crows = calls[cid]
        chunks = sorted(cd.glob("chunk_*.wav"))
        keys = [f"{cid_str}/{ch.name}" for ch in chunks]

        todo = [(k, ch) for k, ch in zip(keys, chunks) if k not in cache]
        if todo:
            def work(item):
                key, ch = item
                asr = make_client()
                t0 = time.time()
                try:
                    out = asr.transcribe_full(ch, context="", language="Hindi")
                except Exception as e:
                    out = {"transcript": "", "raw_text": "", "romanized_text": ""}
                    print(f"  ERROR {key}: {type(e).__name__}: {e}", flush=True)
                return key, out, round(time.time() - t0, 2)

            with ThreadPoolExecutor(max_workers=min(args.workers, len(todo))) as pool:
                futures = [pool.submit(work, item) for item in todo]
                for fut in as_completed(futures):
                    key, out, elapsed = fut.result()
                    entry = {"key": key, "call_id": cid_str, **out, "elapsed": elapsed}
                    append_cache(entry, cache_file)
                    cache[key] = entry

        au = next((r[idx["actual_urdu_transcript"]] for r in crows
                   if isinstance(r[idx["actual_urdu_transcript"]], str)
                   and r[idx["actual_urdu_transcript"]].strip()), "")
        br = next((r[idx["benchmark_roman_urdu"]] for r in crows
                   if isinstance(r[idx["benchmark_roman_urdu"]], str)
                   and r[idx["benchmark_roman_urdu"]].strip()), "")

        raw_texts, roman_texts = [], []
        for i, (ch, key) in enumerate(zip(chunks, keys)):
            c = cache[key]
            raw = c.get("raw_text", "")
            roman = c.get("romanized_text", "")
            v22ph = H.transliterate(raw) if raw else ""
            raw_texts.append(raw)
            roman_texts.append(roman)
            s.append([ch.name, cid, i, au if i == 0 else "", br if i == 0 else "",
                      raw, roman, v22ph])

        if br:
            full_raw = " ".join(raw_texts)
            full_roman = " ".join(roman_texts)
            full_v22ph = H.transliterate(full_raw) if full_raw else ""
            d_roman = diff_words(br, full_roman)
            d_v22ph = diff_words(br, full_v22ph)
            agg["server_roman"][0] += d_roman.matched; agg["server_roman"][1] += d_roman.total
            agg["v22ph"][0] += d_v22ph.matched; agg["v22ph"][1] += d_v22ph.total
            n_scored += 1
            s2.append([cid, br, full_roman, round(d_roman.accuracy, 2), werp(*wer_acc(br, full_roman)),
                       full_v22ph, round(d_v22ph.accuracy, 2), werp(*wer_acc(br, full_v22ph))])

        t_save = time.time()
        wb.save(str(OUT))
        acc_note = ""
        if agg["v22ph"][1]:
            acc_note = f"  running v22ph acc: {100*agg['v22ph'][0]/agg['v22ph'][1]:.2f}%"
        print(f"  [{ci}/{len(ordered_cids)}] {cid_str[:55]}  "
              f"({len(chunks)} chunks, saved in {time.time()-t_save:.1f}s){acc_note}", flush=True)

    print()
    print("=" * 80)
    print(f"  DYNAMIC CHUNKING — {VARIANT}  ({n_scored} calls scored)")
    print("=" * 80)
    if agg["server_roman"][1]:
        print(f"  server romanized_text (as-is):     {100*agg['server_roman'][0]/agg['server_roman'][1]:.2f}%")
    if agg["v22ph"][1]:
        print(f"  v22_phonetic (local @0.90):        {100*agg['v22ph'][0]/agg['v22ph'][1]:.2f}%")

    s2.append([])
    s2.append(["CORPUS (diff_words %)", "",
               "", round(100*agg["server_roman"][0]/agg["server_roman"][1], 2) if agg["server_roman"][1] else 0, "",
               "", round(100*agg["v22ph"][0]/agg["v22ph"][1], 2) if agg["v22ph"][1] else 0, ""])
    wb.save(str(OUT))
    print(f"\n  final save done: {OUT.name}")
    print(f"    - {SHEET_NAME}")
    print(f"    - {SUMMARY_NAME}")


def main():
    ap = argparse.ArgumentParser(description="Benchmark dynamic (silence-aware) chunking via HF /chughtai endpoint")
    sub = ap.add_subparsers(dest="command", required=True)

    def add_common(p):
        p.add_argument("--endpoint", default="hf", choices=["hf", "gpu"],
                        help="hf = HF Space (default); gpu = local GPU box (192.168.99.117:7873)")
        p.add_argument("--calls", type=int, default=0, help="Limit to N calls (0 = all)")
        p.add_argument("--workers", type=int, default=4, help="Concurrent connections per call")

    p_resume = sub.add_parser("resume", help="Resume from cache (skips already-transcribed chunks)")
    add_common(p_resume)
    p_redo = sub.add_parser("redo", help="Clear cache and start fresh")
    add_common(p_redo)

    args = ap.parse_args()
    args.restart = (args.command == "redo")

    _kill_previous()
    atexit.register(_cleanup_pid)
    run(args)


if __name__ == "__main__":
    main()
