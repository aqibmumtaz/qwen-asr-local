"""
Two-pass acoustic contextual biasing benchmark — LOCAL or REMOTE Qwen3-ASR.

Pipeline per chunk (two-pass):
  Pass 1: Audio → Qwen3-ASR(context="", language="Hindi") → rough Hindi
  Retrieve: phonetic encoder picks top-k entity names relevant to pass 1
  Pass 2: Audio → Qwen3-ASR(context=retrieved_names, language="Hindi") → biased Hindi
  Transliterate: hindi_to_roman_urdu → Roman Urdu
  Score: diff_words + WER vs benchmark_roman_urdu

Backend modes:
  local   Qwen3-ASR loaded locally (--device cpu/mps). Proper context= biasing.
  remote  Remote /en WebSocket endpoint. Context passed as session instructions
          (server may not wire it to model context= — verify on your deployment).

Compares per call:
  prev       vendor's model_output_hindi → transliterate (no re-ASR)
  pass1      re-ASR, NO context (baseline)
  pass2      re-ASR, WITH retrieved entity context (acoustic contextual biasing)

Resumable: cached to transcription_cache_acoustic_biasing.jsonl.

  python benchmark/benchmark_acoustic_biasing.py --transcribe --calls 1 --backend local
  python benchmark/benchmark_acoustic_biasing.py --transcribe --calls 80 --backend remote
  python benchmark/benchmark_acoustic_biasing.py --report
"""
from __future__ import annotations

import argparse
import importlib
import json
import os
import re
import sys
import time
import warnings
from collections import defaultdict
from difflib import SequenceMatcher
from pathlib import Path

warnings.filterwarnings("ignore")
HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
sys.path.insert(0, str(ROOT)); sys.path.insert(0, str(HERE))

import signal
import openpyxl
from test_accuracy import diff_words

XLSX = HERE / "lab_test_80_calls_urdu_roman_urdu.xlsx"
AUDIO = HERE / "lab_test_80_audios_chunks_25s"
CACHE = HERE / "transcription_cache_acoustic_biasing.jsonl"
PIDFILE = HERE / ".benchmark_acoustic_biasing.pid"
MODEL_ID = "Qwen/Qwen3-ASR-1.7B"
CAP = re.compile(r"\b[A-Z][a-z]{2,}\b")


def _kill_previous():
    """Kill any previous benchmark process and take over."""
    if PIDFILE.exists():
        try:
            old_pid = int(PIDFILE.read_text().strip())
            if old_pid != os.getpid():
                os.kill(old_pid, signal.SIGTERM)
                print(f"Killed previous benchmark (PID {old_pid})", flush=True)
        except (ProcessLookupError, ValueError, PermissionError):
            pass  # already dead or invalid
    PIDFILE.write_text(str(os.getpid()))


def _cleanup_pid():
    """Remove PID file on exit."""
    try:
        if PIDFILE.exists() and PIDFILE.read_text().strip() == str(os.getpid()):
            PIDFILE.unlink()
    except Exception:
        pass


def cache_path_for(backend: str) -> Path:
    """Return a per-backend cache file so runs don't mix."""
    if backend in ("gpu-remote",):
        return HERE / f"transcription_cache_acoustic_biasing_{backend}.jsonl"
    return CACHE   # local / remote share the original cache


# ── data loading ─────────────────────────────────────────────────────────────

def load_calls():
    wb = openpyxl.load_workbook(str(XLSX)); ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    idx = {h: i for i, h in enumerate(rows[0])}
    calls = defaultdict(list)
    for r in rows[1:]:
        calls[r[idx["call_id"]]].append(r)
    for cid in calls:
        calls[cid].sort(key=lambda x: x[idx["chunk_index"]] or 0)
    return calls, idx


# ── cache ────────────────────────────────────────────────────────────────────

def read_cache(cache_file: Path = CACHE) -> dict:
    out = {}
    if cache_file.exists():
        for line in cache_file.read_text(encoding="utf-8").splitlines():
            if line.strip():
                r = json.loads(line)
                out[r["key"]] = r
    return out


def append_cache(entry: dict, cache_file: Path = CACHE):
    with cache_file.open("a", encoding="utf-8") as f:
        f.write(json.dumps(entry, ensure_ascii=False) + "\n")


# ── ASR backends ─────────────────────────────────────────────────────────────

class LocalASR:
    """Local Qwen3-ASR with proper context= biasing."""

    def __init__(self, device: str = "cpu"):
        import torch
        from qwen_asr import Qwen3ASRModel

        self.device = device
        print(f"[local] Loading {MODEL_ID} on {device} ...", flush=True)
        t0 = time.time()
        self._model = Qwen3ASRModel.from_pretrained(
            MODEL_ID, dtype=torch.float32, device_map=device,
            max_new_tokens=256)
        print(f"[local] Loaded in {time.time()-t0:.0f}s", flush=True)

    def transcribe(self, audio, context: str = "", language: str = "Hindi") -> str:
        out = self._model.transcribe(
            audio=[str(audio)], context=[context], language=[language])
        return out[0].text if out else ""


class RemoteASR:
    """Remote /en WebSocket — context passed as session instructions."""

    def __init__(self):
        from acoustic_contextual_biasing.asr import RemoteASR as _R
        self._asr = _R(variant="en")
        print("[remote] Using /en endpoint", flush=True)

    def transcribe(self, audio, context: str = "", language: str = "Hindi") -> str:
        return self._asr.transcribe(audio, context=context, language=language)


class GpuRemoteASR:
    """GPU server (gpu_remote_asr.py) — proper context= biasing via WebSocket."""

    def __init__(self, url: str = "ws://192.168.99.117:8910"):
        from acoustic_contextual_biasing.gpu_remote_asr import GpuRemoteASR as _G
        self._asr = _G(url=url)

    def transcribe(self, audio, context: str = "", language: str = "Hindi") -> str:
        return self._asr.transcribe(audio, context=context, language=language)


def make_asr(backend: str, device: str = "cpu", gpu_url: str = "ws://192.168.99.117:8910"):
    if backend == "local":
        return LocalASR(device=device)
    if backend == "remote":
        return RemoteASR()
    if backend == "gpu-remote":
        return GpuRemoteASR(url=gpu_url)
    raise ValueError(f"unknown backend: {backend}")


# ── transcribe (two-pass) ───────────────────────────────────────────────────

def transcribe(args):
    import importlib as _imp
    os.environ["LEXICON"] = "v22"; os.environ["RESOLVER"] = "0"
    os.environ["PHONETIC"] = "1"; os.environ["PHONETIC_THRESHOLD"] = "0.9"
    import hindi_to_roman_urdu as _H
    _imp.reload(_H)

    asr = make_asr(args.backend, device=args.device, gpu_url=args.gpu_url)

    from acoustic_contextual_biasing.retriever import NameRetriever
    retr_device = args.device if args.backend == "local" else "cpu"
    retriever = NameRetriever(device=retr_device)
    print(f"Retriever loaded ({len(retriever.names)} gazetteer names)", flush=True)

    cfile = cache_path_for(args.backend)
    if args.restart and cfile.exists():
        cfile.unlink()
        print(f"Cache cleared: {cfile.name}  (restart mode)", flush=True)
    calls, _ = load_calls()
    cache = read_cache(cfile)

    # Build todo list
    todo = []
    for cid in calls:
        if not (AUDIO / str(cid)).exists():
            continue
        for ch in sorted((AUDIO / str(cid)).glob("chunk_*.wav")):
            key = f"{cid}/{ch.name}"
            if key not in cache:
                todo.append((key, cid, ch))

    # Limit to N calls if requested
    if args.calls > 0:
        seen_calls = set()
        limited = []
        for key, cid, ch in todo:
            seen_calls.add(cid)
            if len(seen_calls) > args.calls:
                break
            limited.append((key, cid, ch))
        todo = limited

    # Count calls and chunks for progress
    todo_calls = sorted(set(cid for _, cid, _ in todo))
    total_calls = len(todo_calls)
    print(f"Chunks to transcribe: {len(todo)} across {total_calls} calls  "
          f"(cache has {len(cache)} chunks)", flush=True)

    current_call = None
    call_num = 0
    call_start = None

    for n, (key, cid, ch) in enumerate(todo, 1):
        # Log call-level progress + update sheet when a call completes
        if cid != current_call:
            if current_call is not None:
                call_elapsed = time.time() - call_start
                print(f"  ── call {call_num}/{total_calls} done in {call_elapsed:.1f}s  "
                      f"(updating sheet...)", flush=True)
                report(args)
                print()
            current_call = cid
            call_num += 1
            call_start = time.time()
            print(f"[Call {call_num}/{total_calls}] {cid[:50]}", flush=True)

        t0 = time.time()

        # Pass 1: no context
        t_p1 = time.time()
        pass1_hindi = asr.transcribe(ch, context="", language="Hindi")
        elapsed_pass1 = round(time.time() - t_p1, 2)

        # Transliterate pass1 to Roman Urdu for retriever (needs Latin text)
        pass1_roman = _H.transliterate(pass1_hindi) if pass1_hindi else ""

        # Retrieve relevant names from pass 1 (romanized), deduplicate
        raw_names = retriever.retrieve(pass1_roman, k=args.k * 2)
        seen_lower = set()
        names = []
        for n in raw_names:
            if n.lower() not in seen_lower:
                seen_lower.add(n.lower())
                names.append(n)
            if len(names) >= args.k:
                break
        ctx = ", ".join(names)

        # Pass 2: with retrieved names as context
        t_p2 = time.time()
        pass2_hindi = asr.transcribe(ch, context=ctx, language="Hindi")
        elapsed_pass2 = round(time.time() - t_p2, 2)

        entry = {
            "key": key,
            "pass1": pass1_hindi,
            "pass2": pass2_hindi,
            "context": ctx,
            "names": names,
            "backend": args.backend,
            "elapsed_pass1": elapsed_pass1,
            "elapsed_pass2": elapsed_pass2,
        }
        append_cache(entry, cfile)
        elapsed = time.time() - t0
        print(f"    {ch.name}  pass1={elapsed_pass1:.1f}s  pass2={elapsed_pass2:.1f}s  "
              f"total={elapsed:.0f}s  names={names[:4]}",
              flush=True)

    # Log final call + update sheet
    if current_call is not None:
        call_elapsed = time.time() - call_start
        print(f"  ── call {call_num}/{total_calls} done in {call_elapsed:.1f}s  "
              f"(updating sheet...)", flush=True)

    # Final sheet update
    print(f"\nTranscription complete. {len(todo)} chunks across {total_calls} calls.",
          flush=True)
    report(args)


# ── scoring helpers ──────────────────────────────────────────────────────────

def fuzzy_in(name, words, thr=0.75):
    n = name.lower()
    return any(SequenceMatcher(a=n, b=w, autojunk=False).ratio() >= thr
               for w in words)


def wer_acc(benchmark: str, hypothesis: str) -> tuple:
    from test_accuracy import normalize_tokens
    ref = normalize_tokens(benchmark)
    hyp = normalize_tokens(hypothesis)
    if not ref:
        return (0, 0)
    dp = list(range(len(hyp) + 1))
    for rw in ref:
        nd = [dp[0] + 1] + [0] * len(hyp)
        for j, hw in enumerate(hyp):
            nd[j + 1] = min(dp[j] + (0 if hw == rw else 1),
                            dp[j + 1] + 1, nd[j] + 1)
        dp = nd
    return (dp[len(hyp)], len(ref))


# ── report + Excel ───────────────────────────────────────────────────────────

def report(args):
    # Load transliterator WITHOUT phonetic correction (plain)
    os.environ["LEXICON"] = "v22"; os.environ["RESOLVER"] = "0"
    os.environ["PHONETIC"] = "0"
    import hindi_to_roman_urdu as H
    importlib.reload(H)

    # Load phonetic corrector separately for +pc variants
    from phonetic_contrastive_model.corrector import PhoneticContrastiveCorrector
    pc = PhoneticContrastiveCorrector.load(threshold=0.90)

    def translit(hindi: str) -> str:
        """Plain transliteration (no phonetic correction)."""
        return H.transliterate(hindi) if hindi else ""

    def translit_pc(hindi: str) -> str:
        """Transliteration + phonetic contrastive correction."""
        plain = H.transliterate(hindi) if hindi else ""
        return pc.resolve_text(plain) if plain else ""

    cfile = cache_path_for(args.backend)
    calls, idx = load_calls()
    cache = read_cache(cfile)

    # 6 scoring configs: 3 base × {plain, +phonetic_correction}
    cfgs = ["prev", "prev_pc", "pass1", "pass1_pc", "pass2", "pass2_pc"]
    agg = {c: [0, 0] for c in cfgs}
    nm = {c: [0, 0] for c in cfgs}
    n_calls = n_full = 0

    per_call_results = {}
    all_chunk_rows = []

    for cid, cr in calls.items():
        bench = next((x[idx["benchmark_roman_urdu"]] for x in cr
                      if isinstance(x[idx["benchmark_roman_urdu"]], str)
                      and x[idx["benchmark_roman_urdu"]].strip()), None)
        actual_urdu = ""
        if "actual_urdu_transcript" in idx:
            actual_urdu = next(
                (x[idx["actual_urdu_transcript"]] for x in cr
                 if isinstance(x[idx["actual_urdu_transcript"]], str)
                 and x[idx["actual_urdu_transcript"]].strip()), "")
        if not bench or not (AUDIO / str(cid)).exists():
            continue
        n_calls += 1
        chunks = sorted((AUDIO / str(cid)).glob("chunk_*.wav"))
        keys = [f"{cid}/{ch.name}" for ch in chunks]
        if not all(k in cache for k in keys):
            continue
        n_full += 1
        gnames = sorted(set(CAP.findall(bench)))

        # Hindi texts for the 3 base configs
        hindi_texts = {
            "prev": " ".join(
                str(x[idx["model_output_hindi"]]).strip() for x in cr
                if isinstance(x[idx["model_output_hindi"]], str)
                and x[idx["model_output_hindi"]].strip()),
            "pass1": " ".join(cache[k]["pass1"] for k in keys),
            "pass2": " ".join(cache[k]["pass2"] for k in keys),
        }

        for i, (ch, key) in enumerate(zip(chunks, keys)):
            cr_row = cr[i] if i < len(cr) else cr[0]
            hindi_prev = (str(cr_row[idx["model_output_hindi"]]).strip()
                          if isinstance(cr_row[idx["model_output_hindi"]], str)
                          else "")
            prev_roman = (str(cr_row[idx["model_output_roman_urdu"]]).strip()
                          if isinstance(cr_row[idx["model_output_roman_urdu"]], str)
                          else "")
            c = cache[key]
            p1_plain = translit(c["pass1"])
            p2_plain = translit(c["pass2"])
            row = {
                "audio_name": ch.name,
                "call_id": cid,
                "chunk_index": i,
                "actual_urdu_transcript": actual_urdu if i == 0 else "",
                "benchmark_roman_urdu": bench if i == 0 else "",
                "model_output_hindi": hindi_prev,
                "model_output_roman_urdu": prev_roman,
                "pass1_hindi": c["pass1"],
                "pass1_roman_urdu": p1_plain,
                "pass1_roman_urdu_pc": pc.resolve_text(p1_plain) if p1_plain else "",
                "elapsed_pass1": c.get("elapsed_pass1", ""),
                "context_biasing_list": c.get("context", ""),
                "pass2_hindi": c["pass2"],
                "model_output_acoustic_biasing": p2_plain,
                "model_output_acoustic_biasing_pc": pc.resolve_text(p2_plain) if p2_plain else "",
                "elapsed_pass2": c.get("elapsed_pass2", ""),
            }
            all_chunk_rows.append(row)

        # Score all 6 configs
        d_call = {"bench": bench}
        for base in ["prev", "pass1", "pass2"]:
            # Plain (no phonetic correction)
            roman = translit(hindi_texts[base])
            d = diff_words(bench, roman)
            agg[base][0] += d.matched; agg[base][1] += d.total
            words = roman.lower().split()
            nm[base][0] += sum(1 for n in gnames if fuzzy_in(n, words))
            nm[base][1] += len(gnames)
            d_call[base] = {"roman": roman, "acc": d, "wer": wer_acc(bench, roman)}

            # +phonetic correction
            pc_key = f"{base}_pc"
            roman_pc = pc.resolve_text(roman) if roman else ""
            d_pc = diff_words(bench, roman_pc)
            agg[pc_key][0] += d_pc.matched; agg[pc_key][1] += d_pc.total
            words_pc = roman_pc.lower().split()
            nm[pc_key][0] += sum(1 for n in gnames if fuzzy_in(n, words_pc))
            nm[pc_key][1] += len(gnames)
            d_call[pc_key] = {"roman": roman_pc, "acc": d_pc, "wer": wer_acc(bench, roman_pc)}

        per_call_results[cid] = d_call

    label = {
        "prev":     "vendor hindi",
        "prev_pc":  "vendor hindi + phonetic correction",
        "pass1":    "pass 1 — no context",
        "pass1_pc": "pass 1 — no context + phonetic correction",
        "pass2":    "pass 2 — acoustic biasing",
        "pass2_pc": "pass 2 — acoustic biasing + phonetic correction",
    }
    print("=" * 80)
    print(f"  ACOUSTIC BIASING + PHONETIC CORRECTION — {n_full}/{n_calls} calls")
    print("=" * 80)
    print(f"  {'config':<52}{'diff_words':>11}{'name recovery':>16}")
    print("  " + "-" * 78)
    for c in cfgs:
        if agg[c][1] == 0:
            continue
        dw = 100 * agg[c][0] / agg[c][1]
        nr = 100 * nm[c][0] / max(nm[c][1], 1)
        print(f"  {label[c]:<52}{dw:>10.2f}%{nr:>13.1f}%  "
              f"({nm[c][0]}/{nm[c][1]})")
    print()
    # Gains summary
    if agg["pass1"][1] and agg["pass2"][1]:
        g_bias = (100 * agg["pass2"][0] / agg["pass2"][1]
                  - 100 * agg["pass1"][0] / agg["pass1"][1])
        print(f"  biasing gain (pass2 - pass1):           {g_bias:+.2f} pts")
    if agg["pass1"][1] and agg["pass1_pc"][1]:
        g_pc = (100 * agg["pass1_pc"][0] / agg["pass1_pc"][1]
                - 100 * agg["pass1"][0] / agg["pass1"][1])
        print(f"  phonetic gain (pass1_pc - pass1):       {g_pc:+.2f} pts")
    if agg["pass1"][1] and agg["pass2_pc"][1]:
        g_both = (100 * agg["pass2_pc"][0] / agg["pass2_pc"][1]
                  - 100 * agg["pass1"][0] / agg["pass1"][1])
        print(f"  combined gain (pass2_pc - pass1):       {g_both:+.2f} pts")
    print()

    if not per_call_results:
        print("  (no complete calls to write to Excel)")
        return

    # Use the CLI backend arg as sheet tag (caches are per-backend now)
    backend_tag = args.backend

    OUT = HERE / "lab_test_80_calls_urdu_roman_urdu_benchmarked.xlsx"
    wb_out = openpyxl.load_workbook(str(OUT))

    # --- on restart, delete all old sheets for this backend ---
    if args.restart:
        for old_sheet in [f"acoustic_biasing_per_chunk_{backend_tag}",
                          f"acoustic_biasing_per_call_{backend_tag}",
                          f"acoustic_biasing_summary_{backend_tag}"]:
            if old_sheet in wb_out.sheetnames:
                del wb_out[old_sheet]
        print(f"  Old sheets for '{backend_tag}' deleted (restart mode)", flush=True)

    # --- per-chunk sheet ---
    chunk_sheet = f"acoustic_biasing_per_chunk_{backend_tag}"
    if chunk_sheet in wb_out.sheetnames:
        del wb_out[chunk_sheet]
    s2 = wb_out.create_sheet(chunk_sheet)
    headers = ["audio_name", "call_id", "chunk_index",
               "actual_urdu_transcript", "benchmark_roman_urdu",
               "model_output_hindi", "model_output_roman_urdu",
               "pass1_hindi", "pass1_roman_urdu", "pass1_roman_urdu_pc",
               "elapsed_pass1",
               "context_biasing_list",
               "pass2_hindi", "model_output_acoustic_biasing",
               "model_output_acoustic_biasing_pc", "elapsed_pass2"]
    s2.append(headers)
    for row in all_chunk_rows:
        s2.append([row.get(h, "") for h in headers])

    # --- per-call sheet ---
    call_sheet = f"acoustic_biasing_per_call_{backend_tag}"
    if call_sheet in wb_out.sheetnames:
        del wb_out[call_sheet]
    s3 = wb_out.create_sheet(call_sheet)
    s3.append(["call_id", "benchmark_roman_urdu",
               "out_prev", "acc_prev", "wer_prev",
               "out_prev_pc", "acc_prev_pc", "wer_prev_pc",
               "out_pass1", "acc_pass1", "wer_pass1",
               "out_pass1_pc", "acc_pass1_pc", "wer_pass1_pc",
               "out_pass2", "acc_pass2", "wer_pass2",
               "out_pass2_pc", "acc_pass2_pc", "wer_pass2_pc"])

    def werp(e, n):
        return round(100 * max(0.0, 1 - e / n), 2) if n else None

    for cid, d in per_call_results.items():
        row = [cid, d["bench"]]
        for c in cfgs:
            row.append(d[c]["roman"])
            row.append(round(d[c]["acc"].accuracy, 4))
            row.append(werp(*d[c]["wer"]))
            row.append(werp(*d[c]["wer"]))
        s3.append(row)

    s3.append([])
    corpus_row = ["CORPUS (diff_words %)", ""]
    for c in cfgs:
        dw = 100 * agg[c][0] / agg[c][1] if agg[c][1] else 0
        corpus_row += [round(dw, 2), "", ""]
    s3.append(corpus_row)

    # --- summary sheet ---
    summary_sheet = f"acoustic_biasing_summary_{backend_tag}"
    if summary_sheet in wb_out.sheetnames:
        del wb_out[summary_sheet]
    s4 = wb_out.create_sheet(summary_sheet)
    s4.append(["metric",
               "vendor", "vendor+pc",
               "pass1", "pass1+pc",
               "pass2 (biasing)", "pass2+pc",
               "biasing gain", "phonetic gain", "combined gain"])

    def pct(c):
        return round(100 * agg[c][0] / agg[c][1], 2) if agg[c][1] else 0

    dw = {c: pct(c) for c in cfgs}
    s4.append(["diff_words %",
               dw["prev"], dw["prev_pc"],
               dw["pass1"], dw["pass1_pc"],
               dw["pass2"], dw["pass2_pc"],
               round(dw["pass2"] - dw["pass1"], 2),
               round(dw["pass1_pc"] - dw["pass1"], 2),
               round(dw["pass2_pc"] - dw["pass1"], 2)])

    def nr_pct(c):
        return round(100 * nm[c][0] / max(nm[c][1], 1), 2)

    nr = {c: nr_pct(c) for c in cfgs}
    s4.append(["name recovery %",
               nr["prev"], nr["prev_pc"],
               nr["pass1"], nr["pass1_pc"],
               nr["pass2"], nr["pass2_pc"],
               round(nr["pass2"] - nr["pass1"], 2),
               round(nr["pass1_pc"] - nr["pass1"], 2),
               round(nr["pass2_pc"] - nr["pass1"], 2)])

    s4.append([])
    s4.append(["backend", backend_tag])
    s4.append(["run mode", "restart (fresh)" if args.restart else "resume"])
    s4.append(["calls scored", n_full])
    s4.append(["total calls", n_calls])
    s4.append(["names in gazetteer", nm["prev"][1]])

    # Also remove old sheet names if they exist (migration)
    for old in ["model_acoustic_biasing", "benchmark_acoustic_biasing"]:
        if old in wb_out.sheetnames:
            del wb_out[old]

    wb_out.save(str(OUT))
    print(f"  written to: {OUT.name}")
    print(f"    - {chunk_sheet}  (per-chunk two-pass output)")
    print(f"    - {call_sheet}  (per-call accuracy scores)")
    print(f"    - {summary_sheet}  (corpus-level summary)")
    print()


# ── main ─────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(
        description="Two-pass acoustic contextual biasing benchmark")
    sub = ap.add_subparsers(dest="command", required=True)

    # Shared args for both subcommands
    def add_common(p):
        p.add_argument("--backend", default="local",
                        choices=["local", "remote", "gpu-remote"],
                        help="local = CPU; remote = /en WebSocket; gpu-remote = gpu_remote_asr.py GPU")
        p.add_argument("--calls", type=int, default=0,
                        help="Limit to N calls (0 = all)")
        p.add_argument("--k", type=int, default=15,
                        help="Top-k retrieved names for pass 2 biasing")
        p.add_argument("--device", default="cpu", choices=["cpu", "mps", "cuda"],
                        help="Device for local backend (cpu/mps/cuda)")
        p.add_argument("--gpu-url", default="ws://192.168.99.117:8910",
                        help="URL of gpu_remote_asr.py WebSocket server (for gpu-remote backend)")

    # resume — continue from cache
    p_resume = sub.add_parser("resume",
        help="Resume benchmarking from where you left off (cached chunks are skipped)")
    add_common(p_resume)

    # redo — clear cache + delete old sheets, start from scratch
    p_redo = sub.add_parser("redo",
        help="Redo all benchmarking from scratch (clears cache, deletes old sheets)")
    add_common(p_redo)

    # report — just generate report from existing cache
    p_report = sub.add_parser("report",
        help="Generate report + Excel sheets from existing cache (no transcription)")
    add_common(p_report)

    args = ap.parse_args()
    args.restart = (args.command == "redo")

    # Kill any previous benchmark, register this PID, clean up on exit
    import atexit
    _kill_previous()
    atexit.register(_cleanup_pid)

    if args.command in ("resume", "redo"):
        args.transcribe = True
        transcribe(args)
    elif args.command == "report":
        args.transcribe = False
        report(args)


if __name__ == "__main__":
    main()
