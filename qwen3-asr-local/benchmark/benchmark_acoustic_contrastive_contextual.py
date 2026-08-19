#!/usr/bin/env python3
"""
Benchmark the acoustic_contrastive_contextual_model_finetuned adapter (via
splice_inference.SpliceASR) against the same 72-call corpus every other
pipeline in this repo has been scored against this session.

*** NOT RUN -- needs a trained adapter (GPU), which needs train_lora.py to
have actually completed. Structure mirrors benchmark_chunks_dynamic.py 1:1
per the plan's explicit requirement (same CLI shape, same resumability,
same per-call incremental sheet save) -- do not deviate from that pattern,
copy it, which is what this file does. ***

  python benchmark_acoustic_contrastive_contextual.py resume --adapter ../acoustic_contrastive_contextual_model_finetuned/adapters/run1/phase3 --workers 4
  python benchmark_acoustic_contrastive_contextual.py resume --calls 5    # smoke test first
  python benchmark_acoustic_contrastive_contextual.py redo               # clear cache, restart
"""
from __future__ import annotations

import argparse
import atexit
import json
import os
import signal
import sys
import time
import warnings
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

warnings.filterwarnings("ignore")
HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(HERE))
sys.path.insert(0, str(ROOT / "acoustic_contrastive_contextual_model_finetuned"))

import openpyxl
from test_accuracy import diff_words, normalize_tokens

XLSX = HERE / "lab_test_80_calls_urdu_roman_urdu.xlsx"
OUT = HERE / "lab_test_80_calls_urdu_roman_urdu_benchmarked.xlsx"
AUDIO = HERE / "lab_test_80_audios_chunks_dynamic"
PIDFILE = HERE / ".benchmark_acoustic_contrastive_contextual.pid"

SHEET_NAME = "model_acoustic_contrastive_contextual"
SUMMARY_NAME = "benchmark_summary_acoustic_contrastive_contextual"


def cache_path_for(adapter_name: str) -> Path:
    safe = adapter_name.replace("/", "_")
    return HERE / f"transcription_cache_acoustic_contrastive_contextual_{safe}.jsonl"


def _kill_previous():
    if PIDFILE.exists():
        try:
            old_pid = int(PIDFILE.read_text().strip())
            if old_pid != os.getpid():
                os.kill(old_pid, signal.SIGTERM)
                print(f"Killed previous run (PID {old_pid})", flush=True)
        except (ProcessLookupError, ValueError, PermissionError):
            pass
    PIDFILE.write_text(str(os.getpid()))


def _cleanup_pid():
    try:
        if PIDFILE.exists() and PIDFILE.read_text().strip() == str(os.getpid()):
            PIDFILE.unlink()
    except Exception:
        pass


def load_calls():
    wb = openpyxl.load_workbook(str(XLSX))
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    idx = {h: i for i, h in enumerate(rows[0])}
    calls = defaultdict(list)
    for r in rows[1:]:
        calls[r[idx["call_id"]]].append(r)
    return idx, calls


def read_cache(cache_file: Path) -> dict:
    out = {}
    if cache_file.exists():
        for line in cache_file.read_text(encoding="utf-8").splitlines():
            if line.strip():
                r = json.loads(line)
                out[r["key"]] = r
    return out


def append_cache(entry: dict, cache_file: Path):
    with cache_file.open("a", encoding="utf-8") as f:
        f.write(json.dumps(entry, ensure_ascii=False) + "\n")


def wer_acc(benchmark: str, hypothesis: str) -> tuple[int, int]:
    ref = normalize_tokens(benchmark)
    hyp = normalize_tokens(hypothesis)
    if not ref:
        return (0, 0)
    dp = list(range(len(hyp) + 1))
    for rw in ref:
        nd = [dp[0] + 1] + [0] * len(hyp)
        for j, hw in enumerate(hyp):
            nd[j + 1] = min(dp[j] + (0 if hw == rw else 1), dp[j + 1] + 1, nd[j] + 1)
        dp = nd
    return (dp[len(hyp)], len(ref))


def werp(e, n):
    return round(100 * max(0.0, 1 - e / n), 2) if n else None


def run(args):
    from splice_inference import SpliceASR

    os.environ["LEXICON"] = "v22"; os.environ["RESOLVER"] = "0"
    os.environ["PHONETIC"] = "1"; os.environ["PHONETIC_THRESHOLD"] = "0.9"
    import importlib
    import hindi_to_roman_urdu as H
    importlib.reload(H)

    cache_file = cache_path_for(args.adapter)
    if args.restart and cache_file.exists():
        cache_file.unlink()
        print("Cache cleared (redo mode)", flush=True)

    idx, calls = load_calls()
    cache = read_cache(cache_file)
    call_dirs = {d.name: d for d in AUDIO.iterdir() if d.is_dir()}

    print(f"Loading SpliceASR (adapter: {args.adapter})...", flush=True)
    asr = SpliceASR(adapter_path=args.adapter)
    print("  loaded.", flush=True)

    ordered_cids = [cid for cid in calls if str(cid) in call_dirs]
    if args.calls > 0:
        ordered_cids = ordered_cids[:args.calls]
    print(f"Calls to process: {len(ordered_cids)}", flush=True)

    print(f"Loading workbook ({OUT.name}) once...", flush=True)
    wb = openpyxl.load_workbook(str(OUT))
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    s = wb.create_sheet(SHEET_NAME)
    headers = ["audio_name", "call_id", "chunk_index", "actual_urdu_transcript",
               "benchmark_roman_urdu", "model_output_hindi",
               "model_output_roman_urdu", "model_output_v22_phonetic"]
    s.append(headers)

    if SUMMARY_NAME in wb.sheetnames:
        del wb[SUMMARY_NAME]
    s2 = wb.create_sheet(SUMMARY_NAME)
    s2.append(["call_id", "benchmark_roman_urdu", "out", "acc", "wer"])
    wb.save(str(OUT))
    print("  sheets created, initial save done.", flush=True)

    agg = [0, 0]
    n_scored = 0

    for ci, cid in enumerate(ordered_cids, 1):
        cid_str = str(cid)
        cd = call_dirs[cid_str]
        crows = calls[cid]
        chunks = sorted(cd.glob("chunk_*.wav"))
        keys = [f"{cid_str}/{ch.name}" for ch in chunks]

        for key, ch in zip(keys, chunks):
            if key in cache:
                continue
            t0 = time.time()
            try:
                # SpliceASR.transcribe() decodes in Hindi (per training
                # targets) -- transliterate to Roman the same way every
                # other benchmark script in this repo does it.
                text = asr.transcribe(str(ch))
            except Exception as e:
                text = ""
                print(f"  ERROR {key}: {type(e).__name__}: {e}", flush=True)
            append_cache({"key": key, "call_id": cid_str, "hindi": text,
                         "elapsed": round(time.time() - t0, 2)}, cache_file)
            cache[key] = {"hindi": text}

        au = next((r[idx["actual_urdu_transcript"]] for r in crows
                   if isinstance(r[idx["actual_urdu_transcript"]], str)
                   and r[idx["actual_urdu_transcript"]].strip()), "")
        br = next((r[idx["benchmark_roman_urdu"]] for r in crows
                   if isinstance(r[idx["benchmark_roman_urdu"]], str)
                   and r[idx["benchmark_roman_urdu"]].strip()), "")

        roman_texts = []
        for i, (ch, key) in enumerate(zip(chunks, keys)):
            hindi = cache[key].get("hindi", "")
            roman = H.transliterate(hindi) if hindi else ""
            roman_texts.append(roman)
            s.append([ch.name, cid, i, au if i == 0 else "", br if i == 0 else "",
                      hindi, roman, roman])   # model_output_roman_urdu and
                                                # model_output_v22_phonetic are
                                                # the same value here -- the
                                                # v2.2+phonetic correction is
                                                # already applied by
                                                # H.transliterate() (PHONETIC=1
                                                # set above), there's no
                                                # separate "raw roman" stage
                                                # the way the vendor-Hindi
                                                # pipelines have one

        if br:
            full_text = " ".join(roman_texts)
            d = diff_words(br, full_text)
            agg[0] += d.matched; agg[1] += d.total
            n_scored += 1
            s2.append([cid, br, full_text, round(d.accuracy, 2), werp(*wer_acc(br, full_text))])

        t_save = time.time()
        wb.save(str(OUT))
        acc_note = f"  running acc: {100*agg[0]/agg[1]:.2f}%" if agg[1] else ""
        print(f"  [{ci}/{len(ordered_cids)}] {cid_str[:55]}  "
              f"(saved in {time.time()-t_save:.1f}s){acc_note}", flush=True)

    print()
    print("=" * 80)
    print(f"  ACOUSTIC CONTRASTIVE CONTEXTUAL MODEL — ({n_scored} calls scored)")
    print("=" * 80)
    if agg[1]:
        print(f"  corpus accuracy: {100*agg[0]/agg[1]:.2f}%")
    print()
    print("  Compare against, in this same workbook:")
    print("    64.92%  vendor Hindi + static chunks + local v2.2/phonetic (benchmark_summary)")
    print("    65.87%  /chughtai HTTP + dynamic chunks, no fine-tune (benchmark_summary_chunks_dynamic)")

    s2.append([])
    s2.append(["CORPUS (diff_words %)", "", "", round(100*agg[0]/agg[1], 2) if agg[1] else 0, ""])
    wb.save(str(OUT))
    print(f"\n  final save done: {OUT.name}")
    print(f"    - {SHEET_NAME}")
    print(f"    - {SUMMARY_NAME}")


def main():
    ap = argparse.ArgumentParser(description="Benchmark the acoustic_contrastive_contextual fine-tuned model")
    sub = ap.add_subparsers(dest="command", required=True)

    def add_common(p):
        p.add_argument("--adapter", required=True, help="path to trained LoRA adapter, e.g. ../acoustic_contrastive_contextual_model_finetuned/adapters/run1/phase3")
        p.add_argument("--calls", type=int, default=0, help="Limit to N calls (0 = all)")

    p_resume = sub.add_parser("resume", help="Resume from cache (skips already-transcribed chunks)")
    add_common(p_resume)
    p_redo = sub.add_parser("redo", help="Clear cache and start fresh")
    add_common(p_redo)

    args = ap.parse_args()
    args.restart = (args.command == "redo")

    _kill_previous()
    atexit.register(_cleanup_pid)
    run(args)


if __name__ == "__main__":
    main()
